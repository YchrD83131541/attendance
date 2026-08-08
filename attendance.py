import datetime
import json
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import hashlib
import hmac
import shutil

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF

import labor
import payroll

# make both names available: `labor` module and `attendance` alias used by
# existing GUI code. Avoid `import attendance` to prevent importing this file.
attendance = labor

# backward-compatibility: expose backend functions at module level so
# other modules importing `attendance` can call `attendance.read_rows()` etc.
try:
    read_rows = attendance.read_rows
    calculate_actual_hours = attendance.calculate_actual_hours
    find_row = attendance.find_row
    checkin = attendance.checkin
    checkout = attendance.checkout
    add_lessons = attendance.add_lessons
    update_record = attendance.update_record
    delete_record = attendance.delete_record
except Exception:

    pass


# ── カラーパレット ────────────────────────────────
C_BG        = "#F0F4FF"   # 全体背景
C_HEADER    = "#4A7FD4"   # ヘッダー帯
C_HEADER_FG = "#FFFFFF"
C_BTN_CHECK = "#4C924C"   # 出勤ボタン
C_BTN_OUT   = "#E87C2B"   # 退勤ボタン
C_BTN_LESS  = "#5BC0DE"   # レッスン追加
C_BTN_SAVE  = "#9B59B6"   # 訂正保存
C_BTN_DEL   = "#E74C3C"   # 削除
C_BTN_REFR  = "#7F8C8D"   # 更新
C_BTN_REP   = "#1D7094"   # 集計
C_BTN_XLS   = "#27AE60"   # Excel出力
C_ROW_ODD   = "#FFFFFF"
C_ROW_EVEN  = "#EAF0FB"
C_SELECT    = "#AED6F1"
C_STATUS_OK = "#1A7A3C"
C_STATUS_NG = "#C0392B"
FONT_MAIN   = ("Meiryo UI", 10)
FONT_BOLD   = ("Meiryo UI", 10, "bold")
FONT_TITLE  = ("Meiryo UI", 13, "bold")

STAFF_FILE      = os.path.join(os.path.expanduser("~"), "Documents", "AttendanceApp", "staff_list.json")
USER_FILE       = os.path.join(os.path.expanduser("~"), "Documents", "AttendanceApp", "users.json")
PAID_LEAVE_FILE = os.path.join(os.path.expanduser("~"), "Documents", "AttendanceApp", "paid_leave.json")

# ── ユーザー認証関連 ────────────────────────────
def _hash_password(password):
    """パスワードをハッシュ化"""
    return hashlib.sha256(password.encode()).hexdigest()

def _load_users():
    """ユーザー情報を読み込む"""
    if os.path.exists(USER_FILE):
        try:
            with open(USER_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_users(users):
    """ユーザー情報を保存"""
    os.makedirs(os.path.dirname(USER_FILE), exist_ok=True)
    with open(USER_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def _verify_user(username, password):
    """ユーザー認証を確認"""
    users = _load_users()
    if username not in users:
        return False
    stored_hash = users[username].get("password_hash", "")
    return hmac.compare_digest(stored_hash, _hash_password(password))

def _create_default_user():
    """デフォルトユーザーを作成"""
    users = _load_users()
    if "admin" not in users:
        users["admin"] = {
            "password_hash": _hash_password("admin"),
            "role": "管理者"
        }
        _save_users(users)
def _load_staff_list():
    if os.path.exists(STAFF_FILE):
        try:
            with open(STAFF_FILE, encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    if all(isinstance(item, str) for item in data):
                        return [{"name": item, "category": "一般"} for item in data]
                    return [item if isinstance(item, dict) else {"name": str(item), "category": "一般"} for item in data]
        except Exception:
            return []
    return []


def _save_staff_list(names):
    os.makedirs(os.path.dirname(STAFF_FILE), exist_ok=True)
    with open(STAFF_FILE, "w", encoding="utf-8") as f:
        json.dump(names, f, ensure_ascii=False, indent=2)


def _load_paid_leave():
    if os.path.exists(PAID_LEAVE_FILE):
        try:
            with open(PAID_LEAVE_FILE, encoding="utf-8") as f:
                data = json.load(f)
                data.setdefault("grants", [])
                data.setdefault("usages", [])
                data.setdefault("requests", [])
                return data
        except Exception:
            return {"grants": [], "usages": [], "requests": []}
    return {"grants": [], "usages": [], "requests": []}


def _get_user_role(username):
    """ユーザーのロール（管理者/ユーザー）を取得"""
    users = _load_users()
    return users.get(username, {}).get("role", "ユーザー")


def _save_paid_leave(data):
    os.makedirs(os.path.dirname(PAID_LEAVE_FILE), exist_ok=True)
    with open(PAID_LEAVE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _next_pl_id(records):
    if not records:
        return 1
    return max(int(r.get("id", 0)) for r in records) + 1


def _calc_paid_leave_balance(name, data, as_of=None):
    """(残日数, 累計付与, 累計取得) を返す。期限切れ付与は除外。"""
    if as_of is None:
        as_of = datetime.date.today().isoformat()
    total_granted = sum(
        float(g.get("days", 0)) for g in data["grants"]
        if g["name"] == name and g.get("expiry_date", "9999-12-31") >= as_of
    )
    total_used = sum(float(u.get("days", 0)) for u in data["usages"] if u["name"] == name)
    remaining = round(total_granted - total_used, 1)
    return remaining, round(total_granted, 1), round(total_used, 1)


def _color_btn(parent, text, command, bg, fg="#FFFFFF", width=10):
    """角丸風の色付きボタン（tk.Button で実装）"""
    return tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=fg, activebackground=bg, activeforeground=fg,
        relief=tk.FLAT, bd=0, padx=10, pady=6,
        font=FONT_BOLD, cursor="hand2", width=width,
    )

# ── ログイン画面 ──────────────────────────────
class LoginWindow(tk.Tk):
    """ログイン画面"""
    def __init__(self, callback=None):
        super().__init__()
        self.title("勤怠管理システム - ログイン")
        self.geometry("350x250")
        self.resizable(False, False)
        self.configure(bg=C_BG)
        self.callback = callback
        self.logged_in_user = None
        
        self._create_widgets()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def _create_widgets(self):
        """ウィジェットを作成"""
        # タイトル
        title_frame = tk.Frame(self, bg=C_HEADER, height=50)
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame, text="🔐 ログイン",
            bg=C_HEADER, fg=C_HEADER_FG, font=FONT_TITLE
        ).pack(pady=12)
        
        # メインフレーム
        main_frame = tk.Frame(self, bg=C_BG)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # ユーザー名
        tk.Label(main_frame, text="ユーザー名:", bg=C_BG, font=FONT_MAIN).pack(anchor=tk.W, pady=(0, 4))
        self.username_var = tk.StringVar()
        username_entry = ttk.Entry(main_frame, textvariable=self.username_var, width=25)
        username_entry.pack(fill=tk.X, pady=(0, 12))
        username_entry.focus()
        
        # パスワード
        tk.Label(main_frame, text="パスワード:", bg=C_BG, font=FONT_MAIN).pack(anchor=tk.W, pady=(0, 4))
        self.password_var = tk.StringVar()
        password_entry = ttk.Entry(main_frame, textvariable=self.password_var, width=25, show="*")
        password_entry.pack(fill=tk.X, pady=(0, 20))
        password_entry.bind("<Return>", lambda e: self.on_login())
        
        # ボタンフレーム
        btn_frame = tk.Frame(main_frame, bg=C_BG)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        _color_btn(btn_frame, "ログイン", self.on_login, C_BTN_CHECK, width=12).pack(side=tk.LEFT, padx=(0, 5))
        _color_btn(btn_frame, "キャンセル", self.on_close, C_BTN_DEL, width=12).pack(side=tk.LEFT)
        
        # メッセージラベル
        self.message_var = tk.StringVar()
        tk.Label(main_frame, textvariable=self.message_var, bg=C_BG, fg=C_STATUS_NG, font=("Meiryo UI", 9)).pack(pady=(15, 0))
    
    def on_login(self):
        """ログイン処理"""
        username = self.username_var.get().strip()
        password = self.password_var.get()
        
        if not username or not password:
            self.message_var.set("ユーザー名とパスワードを入力してください")
            return
        
        if _verify_user(username, password):
            self.logged_in_user = username
            self.destroy()
        else:
            self.message_var.set("ユーザー名またはパスワードが正しくありません")
            self.password_var.set("")
    
    def on_close(self):
        """ウィンドウを閉じる"""
        self.destroy()

#勤怠管理アプリの画面を作るクラス
class AttendanceGUI(tk.Tk):
    def __init__(self, username=None):
        super().__init__() #tkinterのウインドウ機能を有効化
        self.logged_in_user = username
        self.title("スポーツクラブ 勤怠管理")
        self.geometry("1060x580")
        self.configure(bg=C_BG)
        self.sort_by_name = False
        self.staff_list = _load_staff_list()
        self.staff_list.sort(key=lambda item: item.get("name", "") if isinstance(item, dict) else str(item))
        self._edit_original_date = None
        self._overtime_alert_shown = False
        self._apply_style()
        self.create_widgets()
        self.refresh_records()
        # キーボードショートカット
        self.bind("<Control-s>", lambda e: self.on_save_corrections())
        self.bind("<F5>", lambda e: self.refresh_records())
        self.bind("<Control-Return>", lambda e: self.on_register())
        # 起動時に36協定の警告をチェック
        try:
            self.check_overtime_alerts()
        except Exception as exc:
            self.set_status(f"36協定チェックエラー: {exc}", error=True)

    # ── スタイル設定 ──────────────────────────────
    def _apply_style(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        style.configure(".", background=C_BG, font=FONT_MAIN)
        style.configure("TFrame", background=C_BG)
        style.configure("TLabel", background=C_BG, font=FONT_MAIN)
        style.configure("TEntry", font=FONT_MAIN, fieldbackground="#FDFEFF")
        style.configure("TCombobox", font=FONT_MAIN)
        style.configure("TRadiobutton", background=C_BG, font=FONT_MAIN)

        # Treeview
        style.configure(
            "Treeview",
            background=C_ROW_ODD,
            fieldbackground=C_ROW_ODD,
            rowheight=26,
            font=FONT_MAIN,
        )
        style.configure(
            "Treeview.Heading",
            background=C_HEADER,
            foreground=C_HEADER_FG,
            font=FONT_BOLD,
            relief=tk.FLAT,
        )
        style.map("Treeview", background=[("selected", C_SELECT)])
        style.map("Treeview.Heading", background=[("active", "#3A6FC4")])

    # ── ウィジェット構築 ──────────────────────────
    def create_widgets(self):
        # ─ タイトルヘッダー帯
        hdr = tk.Frame(self, bg=C_HEADER, height=44)
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr, text="⚽ スポーツクラブ 勤怠管理システム",
            bg=C_HEADER, fg=C_HEADER_FG, font=FONT_TITLE,
        ).pack(side=tk.LEFT, padx=16, pady=8)
        
        # ログイン情報とログアウトボタン
        if self.logged_in_user:
            user_frame = tk.Frame(hdr, bg=C_HEADER)
            user_frame.pack(side=tk.RIGHT, padx=16, pady=8)
            tk.Label(
                user_frame, text=f"👤 {self.logged_in_user}",
                bg=C_HEADER, fg=C_HEADER_FG, font=FONT_MAIN
            ).pack(side=tk.LEFT, padx=(0, 10))
            _color_btn(
                user_frame, "🚪 ログアウト", self.on_logout, C_BTN_OUT, width=8
            ).pack(side=tk.LEFT)

        # ─ メインフレーム
        frame = tk.Frame(self, bg=C_BG, padx=14, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)

        # ─ ソートボタン行
        sort_row = tk.Frame(frame, bg=C_BG)
        sort_row.pack(fill=tk.X, pady=(0, 4))
        self.sort_button = _color_btn(
            sort_row, "名前順ソート", self.on_toggle_sort, C_BTN_REFR, width=12
        )
        self.sort_button.pack(side=tk.LEFT)
        _color_btn(
            sort_row, "スタッフ管理", self.open_staff_manager, C_BTN_SAVE, width=12
        ).pack(side=tk.LEFT, padx=(6, 0))
        _color_btn(
            sort_row, "ユーザー管理", self.open_user_manager, C_BTN_REP, width=12
        ).pack(side=tk.LEFT, padx=(6, 0))

        # ─ 入力行
        top = tk.Frame(frame, bg=C_BG)
        top.pack(fill=tk.X, pady=4)

        def lbl(parent, text, row, col, padx=0):
            tk.Label(parent, text=text, bg=C_BG, font=FONT_MAIN).grid(
                row=row, column=col, sticky=tk.W, padx=(padx, 2)
            )

        lbl(top, "スタッフ名:", 0, 0)
        self.name_var = tk.StringVar()
        self.name_combo = ttk.Combobox(top, textvariable=self.name_var, width=20, state="readonly")
        self.name_combo.grid(row=0, column=1, sticky=tk.W)
        self._update_staff_combobox()

        lbl(top, "日付:", 0, 2, padx=12)
        self.date_var = tk.StringVar(value=datetime.date.today().isoformat())
        ttk.Entry(top, textvariable=self.date_var, width=14).grid(row=0, column=3, sticky=tk.W)

        lbl(top, "勤務区分:", 0, 4, padx=12)
        self.work_type_var = tk.StringVar(value="通常出勤")
        ttk.Combobox(
            top, textvariable=self.work_type_var,
            values=("通常出勤", "所定休日出勤", "法定休日出勤"),
            width=13, state="readonly",
        ).grid(row=0, column=5, sticky=tk.W)

        # メイン/サブレッスン数・事務作業は time_frame 内（休憩終了の右側）に配置
        self.lessons_main_var  = tk.StringVar(value="0")
        self.lessons_sub_var   = tk.StringVar(value="0")
        self.admin_hours_var   = tk.StringVar(value="0")
        self.admin_minutes_var = tk.StringVar(value="0")
        self.break_var         = tk.StringVar(value="0")

        # ─ 時刻入力・表示行
        time_frame = tk.Frame(frame, bg=C_BG)
        time_frame.pack(fill=tk.X, pady=(0, 6))

        def tlbl(text, row, col, padx=0):
            tk.Label(time_frame, text=text, bg=C_BG, font=FONT_MAIN).grid(
                row=row, column=col, sticky=tk.W, padx=(padx, 2)
            )

        # 入力行（出勤・退勤・休憩開始・休憩終了を一列に）
        tlbl("出勤時間:", 0, 0)
        self.checkin_time_var = tk.StringVar(value="")
        ttk.Entry(time_frame, textvariable=self.checkin_time_var, width=16).grid(row=0, column=1, sticky=tk.W)

        tlbl("退勤時間:", 0, 2, padx=8)
        self.checkout_time_var = tk.StringVar(value="")
        ttk.Entry(time_frame, textvariable=self.checkout_time_var, width=16).grid(row=0, column=3, sticky=tk.W)

        tlbl("休憩時間(分):", 0, 4, padx=8)
        ttk.Entry(time_frame, textvariable=self.break_var, width=6).grid(row=0, column=5, sticky=tk.W)
        self.break_start_var = tk.StringVar(value="")
        self.break_end_var = tk.StringVar(value="")

        _color_btn(
            time_frame, "📋 前回コピー", self.on_copy_last_pattern, C_BTN_REFR, width=11
        ).grid(row=0, column=6, columnspan=2, sticky=tk.W, padx=(8, 0))

        tlbl("メインレッスン数:", 0, 8, padx=8)
        ttk.Entry(time_frame, textvariable=self.lessons_main_var, width=5).grid(row=0, column=9, sticky=tk.W)

        tlbl("サブレッスン数:", 0, 10, padx=8)
        ttk.Entry(time_frame, textvariable=self.lessons_sub_var, width=5).grid(row=0, column=11, sticky=tk.W)

        tlbl("事務作業:", 0, 12, padx=8)
        ttk.Entry(time_frame, textvariable=self.admin_hours_var,   width=4).grid(row=0, column=13, sticky=tk.W)
        tk.Label(time_frame, text="h", bg=C_BG).grid(row=0, column=14, sticky=tk.W)
        ttk.Entry(time_frame, textvariable=self.admin_minutes_var, width=4).grid(row=0, column=15, sticky=tk.W, padx=(4, 0))
        tk.Label(time_frame, text="m", bg=C_BG).grid(row=0, column=16, sticky=tk.W)

        # 表示行（休憩・総勤務・実働・深夜）
        tlbl("休憩:", 1, 0)
        self.break_display_var = tk.StringVar(value="0分")
        tk.Label(time_frame, textvariable=self.break_display_var, bg=C_BG, width=8, anchor=tk.W).grid(row=1, column=1, sticky=tk.W)

        tlbl("総勤務:", 1, 2, padx=8)
        self.total_work_var = tk.StringVar(value="-")
        tk.Label(time_frame, textvariable=self.total_work_var, bg=C_BG, width=10, anchor=tk.W).grid(row=1, column=3, sticky=tk.W)

        tlbl("実働時間:", 1, 4, padx=8)
        self.actual_hours_var = tk.StringVar(value="-")
        tk.Label(time_frame, textvariable=self.actual_hours_var, bg=C_BG, width=10, anchor=tk.W).grid(row=1, column=5, sticky=tk.W)

        tlbl("深夜時間:", 1, 6, padx=8)
        self.night_hours_var = tk.StringVar(value="-")
        tk.Label(time_frame, textvariable=self.night_hours_var, bg=C_BG, width=10, anchor=tk.W).grid(row=1, column=7, sticky=tk.W)

        # ─ ボタン行
        btn_row = tk.Frame(frame, bg=C_BG)
        btn_row.pack(fill=tk.X, pady=(10, 4))

        buttons = [
            ("✅ 登録",       self.on_register,         C_BTN_CHECK),
            ("📚 レッスン追加", self.on_add_lessons,    C_BTN_LESS),
            ("💾 訂正保存",    self.on_save_corrections, C_BTN_SAVE),
            ("🗑 削除",       self.on_delete_selected,  C_BTN_DEL),
            ("📊 集計",       self.show_report,         C_BTN_REP),
            ("📝 36協定管理",  self.open_labor_manager,  "#D35400"),
            ("📥 Excel出力",  self.export_excel,        C_BTN_XLS),
            ("🌴 有給管理",  self.open_paid_leave_manager, "#16A085"),
            ("💰 給与計算",  self.open_payroll_manager,   "#8E44AD"),
            ("🔄 更新",       self.refresh_records,     C_BTN_REFR),
        ]
        for text, cmd, color in buttons:
            _color_btn(btn_row, text, cmd, color, width=11).pack(side=tk.LEFT, padx=3)

        # ─ ステータスバー
        self.status_var = tk.StringVar(value="準備完了")
        self.status_label = tk.Label(
            frame, textvariable=self.status_var,
            bg=C_BG, fg=C_STATUS_OK, font=("Meiryo UI", 9, "italic"), anchor=tk.W
        )
        self.status_label.pack(fill=tk.X, pady=(2, 6))

        # ─ フィルター行
        filter_row = tk.Frame(frame, bg=C_BG)
        filter_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(filter_row, text="月:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        self.filter_month_var = tk.StringVar(value="")
        self.filter_month_combo = ttk.Combobox(filter_row, textvariable=self.filter_month_var,
                                               width=10, state="readonly")
        self.filter_month_combo.pack(side=tk.LEFT, padx=(2, 8))
        tk.Label(filter_row, text="名前:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        self.filter_name_var = tk.StringVar(value="")
        self.filter_name_combo_filter = ttk.Combobox(filter_row, textvariable=self.filter_name_var,
                                                     width=14, state="readonly")
        self.filter_name_combo_filter.pack(side=tk.LEFT, padx=(2, 8))
        _color_btn(filter_row, "絞り込み", self.refresh_records, C_BTN_REFR, width=8).pack(side=tk.LEFT)
        _color_btn(filter_row, "クリア", self.clear_filter, C_BTN_REFR, width=8).pack(side=tk.LEFT, padx=(4, 0))

        # ─ Treeview
        tree_frame = tk.Frame(frame, bg=C_BG)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("date", "name", "work_type", "checkin", "checkout", "break", "total", "lessons", "admin")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15)

        headers = {
            "date":      ("日付",          90,  tk.CENTER),
            "name":      ("スタッフ名",    120, tk.W),
            "work_type": ("勤務区分",      110, tk.CENTER),
            "checkin":   ("出勤",          125, tk.CENTER),
            "checkout":  ("退勤",          125, tk.CENTER),
            "break":     ("休憩",           65, tk.CENTER),
            "total":     ("総勤務",         65, tk.CENTER),
            "lessons":   ("レッスン(M/S)", 120, tk.CENTER),
            "admin":     ("事務作業",      100, tk.CENTER),
        }
        for cid, (htext, w, anchor) in headers.items():
            self.tree.heading(cid, text=htext)
            self.tree.column(cid, width=w, anchor=anchor)

        self.tree.tag_configure("odd",  background=C_ROW_ODD)
        self.tree.tag_configure("even", background=C_ROW_EVEN)

        sb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=sb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

    # ── スタッフ管理 ──────────────────────────────
    def _update_staff_combobox(self):
        self.name_combo["values"] = [s.get("name") for s in self.staff_list if isinstance(s, dict)]

    def _get_staff_category(self, name):
        for staff in self.staff_list:
            if isinstance(staff, dict) and staff.get("name") == name:
                return staff.get("category", "一般")
        return "一般"

    def open_staff_manager(self):
        dlg = tk.Toplevel(self)
        dlg.title("スタッフ管理")
        dlg.geometry("320x440")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        tk.Label(dlg, text="スタッフ一覧", bg=C_BG, font=FONT_BOLD).pack(pady=(10, 4))

        list_frame = tk.Frame(dlg, bg=C_BG)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10)

        lb = tk.Listbox(list_frame, font=FONT_MAIN, selectbackground=C_SELECT, height=12, relief=tk.FLAT, bd=1)
        sb_lb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=lb.yview)
        lb.configure(yscrollcommand=sb_lb.set)
        for staff in self.staff_list:
            if isinstance(staff, dict):
                lb.insert(tk.END, f"{staff.get('name')} ({staff.get('category','一般')})")
            else:
                lb.insert(tk.END, str(staff))
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_lb.pack(side=tk.RIGHT, fill=tk.Y)

        input_frame = tk.Frame(dlg, bg=C_BG)
        input_frame.pack(fill=tk.X, padx=10, pady=(8, 2))
        tk.Label(input_frame, text="名前:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        new_name_var = tk.StringVar()
        entry = ttk.Entry(input_frame, textvariable=new_name_var, width=16)
        entry.pack(side=tk.LEFT, padx=(4, 0))
        tk.Label(input_frame, text="区分:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT, padx=(8, 0))
        new_category_var = tk.StringVar(value="一般")
        ttk.Combobox(input_frame, textvariable=new_category_var,
                     values=("一般", "年少者", "妊産婦"), width=10, state="readonly").pack(side=tk.LEFT, padx=(4, 0))

        def refresh_lb():
            lb.delete(0, tk.END)
            for staff in self.staff_list:
                if isinstance(staff, dict):
                    lb.insert(tk.END, f"{staff.get('name')} ({staff.get('category','一般')})")
                else:
                    lb.insert(tk.END, str(staff))

        def on_select(event=None):
            sel = lb.curselection()
            if not sel:
                return
            item_text = lb.get(sel[0])
            old_name = item_text.split(" (")[0]
            for staff in self.staff_list:
                if isinstance(staff, dict) and staff.get("name") == old_name:
                    new_name_var.set(old_name)
                    new_category_var.set(staff.get("category", "一般"))
                    break

        lb.bind("<<ListboxSelect>>", on_select)

        def add_staff():
            name = new_name_var.get().strip()
            category = new_category_var.get().strip() or "一般"
            if not name:
                return
            if any(staff.get("name") == name for staff in self.staff_list if isinstance(staff, dict)):
                messagebox.showwarning("重複", f"「{name}」は既に登録されています。", parent=dlg)
                return
            self.staff_list.append({"name": name, "category": category})
            self.staff_list.sort(key=lambda item: item.get("name", "") if isinstance(item, dict) else str(item))
            _save_staff_list(self.staff_list)
            self._update_staff_combobox()
            refresh_lb()
            new_name_var.set("")
            entry.focus()

        entry.bind("<Return>", lambda _: add_staff())

        def rename_staff():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("修正", "修正するスタッフをリストから選択してください。", parent=dlg)
                return
            item_text = lb.get(sel[0])
            old_name = item_text.split(" (")[0]
            new_name = new_name_var.get().strip()
            new_cat  = new_category_var.get().strip() or "一般"
            if not new_name:
                messagebox.showwarning("入力エラー", "新しい名前を入力してください。", parent=dlg)
                return
            if new_name != old_name and any(
                s.get("name") == new_name for s in self.staff_list if isinstance(s, dict)
            ):
                messagebox.showwarning("重複", f"「{new_name}」は既に登録されています。", parent=dlg)
                return
            for staff in self.staff_list:
                if isinstance(staff, dict) and staff.get("name") == old_name:
                    staff["name"] = new_name
                    staff["category"] = new_cat
                    break
            self.staff_list.sort(key=lambda item: item.get("name", "") if isinstance(item, dict) else str(item))
            _save_staff_list(self.staff_list)
            if new_name != old_name:
                am = labor.AttendanceManager()
                rows = am.read_rows()
                for r in rows:
                    if r.get("name") == old_name:
                        r["name"] = new_name
                am.save_rows(rows)
                self.refresh_records()
            self._update_staff_combobox()
            refresh_lb()
            self.set_status(f"「{old_name}」→「{new_name}」に修正しました。")

        def delete_staff():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("削除", "削除するスタッフを選択してください。", parent=dlg)
                return
            item_text = lb.get(sel[0])
            name = item_text.split(" (")[0]
            if messagebox.askyesno("確認", f"「{name}」を削除しますか？", parent=dlg):
                self.staff_list = [staff for staff in self.staff_list if not (isinstance(staff, dict) and staff.get("name") == name)]
                _save_staff_list(self.staff_list)
                self._update_staff_combobox()
                refresh_lb()

        def open_payroll_settings():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("給与設定", "対象のスタッフをリストから選択してください。", parent=dlg)
                return
            item_text = lb.get(sel[0])
            name = item_text.split(" (")[0]
            self._open_staff_payroll_dialog(dlg, name, on_saved=refresh_lb)

        btn_frame = tk.Frame(dlg, bg=C_BG)
        btn_frame.pack(pady=(6, 4))
        _color_btn(btn_frame, "追加", add_staff,    C_BTN_CHECK, width=8).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_frame, "修正", rename_staff, C_BTN_SAVE,  width=8).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_frame, "削除", delete_staff, C_BTN_DEL,   width=8).pack(side=tk.LEFT, padx=4)

        btn_frame2 = tk.Frame(dlg, bg=C_BG)
        btn_frame2.pack(pady=(0, 10))
        _color_btn(btn_frame2, "💰 給与設定", open_payroll_settings, "#8E44AD", width=14).pack(side=tk.LEFT, padx=4)

    # ── スタッフ別 給与設定(時給・保険・税区分) ──────────
    def _open_staff_payroll_dialog(self, parent, name, on_saved=None):
        staff = next((s for s in self.staff_list if isinstance(s, dict) and s.get("name") == name), None)
        if staff is None:
            return
        dlg = tk.Toplevel(parent)
        dlg.title(f"給与設定 - {name}")
        dlg.geometry("420x760")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        f = tk.Frame(dlg, bg=C_BG)
        f.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)

        def flbl(text, row):
            tk.Label(f, text=text, bg=C_BG, font=FONT_MAIN).grid(
                row=row, column=0, sticky=tk.W, pady=6)

        r = 0
        flbl("給与形態:", r)
        pay_type_var = tk.StringVar(value=staff.get("pay_type", "時給") or "時給")
        ttk.Combobox(f, textvariable=pay_type_var, values=("時給", "月給"),
                     width=6, state="readonly").grid(row=r, column=1, sticky=tk.W); r += 1

        flbl("時給(円):", r)
        wage_var = tk.StringVar(value=str(staff.get("hourly_wage", 0) or 0))
        ttk.Entry(f, textvariable=wage_var, width=14).grid(row=r, column=1, sticky=tk.W)
        tk.Label(f, text="※給与形態が「時給」の場合に使用", bg=C_BG,
                 font=("Meiryo UI", 8)).grid(row=r + 1, column=0, columnspan=2, sticky=tk.W); r += 2

        flbl("基本給(月額,円):", r)
        base_salary_var = tk.StringVar(value=str(staff.get("monthly_base_salary", 0) or 0))
        ttk.Entry(f, textvariable=base_salary_var, width=14).grid(row=r, column=1, sticky=tk.W)
        tk.Label(f, text="※給与形態が「月給」の場合に使用。固定支給とし、残業・深夜・\n"
                         "休日手当は基本給を所定労働時間(給与計算の料率設定で編集可)で\n"
                         "換算した時間単価により別途加算します。",
                 bg=C_BG, font=("Meiryo UI", 8), justify=tk.LEFT).grid(
            row=r + 1, column=0, columnspan=2, sticky=tk.W); r += 2

        flbl("役職手当(月額,円):", r)
        position_var = tk.StringVar(value=str(staff.get("position_allowance", 0) or 0))
        ttk.Entry(f, textvariable=position_var, width=14).grid(row=r, column=1, sticky=tk.W); r += 1

        flbl("皆勤手当(月額,円):", r)
        attendance_allow_var = tk.StringVar(value=str(staff.get("attendance_allowance", 0) or 0))
        ttk.Entry(f, textvariable=attendance_allow_var, width=14).grid(row=r, column=1, sticky=tk.W)
        tk.Label(f, text="※欠勤等があった月は0円に調整するなど、対象月ごとに見直してください",
                 bg=C_BG, font=("Meiryo UI", 8), justify=tk.LEFT, wraplength=300).grid(
            row=r + 1, column=0, columnspan=2, sticky=tk.W); r += 2

        flbl("住宅手当(月額,円):", r)
        housing_var = tk.StringVar(value=str(staff.get("housing_allowance", 0) or 0))
        ttk.Entry(f, textvariable=housing_var, width=14).grid(row=r, column=1, sticky=tk.W); r += 1

        flbl("交通費(月額,円):", r)
        commute_var = tk.StringVar(value=str(staff.get("commute_allowance", 0) or 0))
        ttk.Entry(f, textvariable=commute_var, width=14).grid(row=r, column=1, sticky=tk.W)
        tk.Label(f, text="※通勤手当は非課税として所得税の課税対象額から除外されます\n"
                         "(社会保険料・雇用保険料の算定基礎には含めます)",
                 bg=C_BG, font=("Meiryo UI", 8), justify=tk.LEFT).grid(
            row=r + 1, column=0, columnspan=2, sticky=tk.W); r += 2

        social_var = tk.BooleanVar(value=bool(staff.get("social_insurance", False)))
        tk.Checkbutton(f, text="社会保険(健康保険・厚生年金)に加入", variable=social_var,
                        bg=C_BG, font=FONT_MAIN, anchor=tk.W).grid(
            row=r, column=0, columnspan=2, sticky=tk.W, pady=4); r += 1

        employment_var = tk.BooleanVar(value=bool(staff.get("employment_insurance", False)))
        tk.Checkbutton(f, text="雇用保険に加入", variable=employment_var,
                        bg=C_BG, font=FONT_MAIN, anchor=tk.W).grid(
            row=r, column=0, columnspan=2, sticky=tk.W, pady=4); r += 1

        flbl("生年月日(YYYY-MM-DD):", r)
        birth_var = tk.StringVar(value=staff.get("birth_date", "") or "")
        ttk.Entry(f, textvariable=birth_var, width=14).grid(row=r, column=1, sticky=tk.W)
        tk.Label(f, text="※40〜64歳は介護保険料が加算されます", bg=C_BG,
                 font=("Meiryo UI", 8)).grid(row=r + 1, column=0, columnspan=2, sticky=tk.W); r += 2

        flbl("扶養親族等の数:", r)
        dependents_var = tk.StringVar(value=str(staff.get("dependents", 0) or 0))
        ttk.Entry(f, textvariable=dependents_var, width=6).grid(row=r, column=1, sticky=tk.W); r += 1

        spouse_var = tk.BooleanVar(value=bool(staff.get("has_spouse_deduction", False)))
        tk.Checkbutton(f, text="源泉控除対象配偶者がいる", variable=spouse_var,
                        bg=C_BG, font=FONT_MAIN, anchor=tk.W).grid(
            row=r, column=0, columnspan=2, sticky=tk.W, pady=4); r += 1

        flbl("所得税区分:", r)
        tax_table_var = tk.StringVar(value=staff.get("tax_table", "甲") or "甲")
        ttk.Combobox(f, textvariable=tax_table_var, values=("甲", "乙"),
                     width=5, state="readonly").grid(row=r, column=1, sticky=tk.W); r += 1

        flbl("住民税(月額,円):", r)
        resident_var = tk.StringVar(value=str(staff.get("resident_tax_monthly", 0) or 0))
        ttk.Entry(f, textvariable=resident_var, width=14).grid(row=r, column=1, sticky=tk.W)
        tk.Label(f, text="※市区町村の特別徴収税額通知書の月額を入力", bg=C_BG,
                 font=("Meiryo UI", 8)).grid(row=r + 1, column=0, columnspan=2, sticky=tk.W); r += 2

        def save():
            try:
                wage = float(wage_var.get().strip() or 0)
                if wage < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "時給は0以上の数値を入力してください。", parent=dlg)
                return
            try:
                base_salary = float(base_salary_var.get().strip() or 0)
                if base_salary < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "基本給は0以上の数値を入力してください。", parent=dlg)
                return
            allowance_fields = [
                ("役職手当", position_var), ("皆勤手当", attendance_allow_var),
                ("住宅手当", housing_var), ("交通費", commute_var),
            ]
            allowance_values = {}
            for label, var in allowance_fields:
                try:
                    v = float(var.get().strip() or 0)
                    if v < 0:
                        raise ValueError
                    allowance_values[label] = v
                except ValueError:
                    messagebox.showwarning("入力エラー", f"{label}は0以上の数値を入力してください。", parent=dlg)
                    return
            birth = birth_var.get().strip()
            if birth:
                try:
                    datetime.date.fromisoformat(birth)
                except ValueError:
                    messagebox.showwarning("入力エラー", "生年月日はYYYY-MM-DD形式で入力してください。", parent=dlg)
                    return
            try:
                dependents = int(dependents_var.get().strip() or 0)
                if dependents < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "扶養親族等の数は0以上の整数を入力してください。", parent=dlg)
                return
            try:
                resident = float(resident_var.get().strip() or 0)
                if resident < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "住民税は0以上の数値を入力してください。", parent=dlg)
                return

            staff["pay_type"]               = pay_type_var.get()
            staff["hourly_wage"]           = wage
            staff["monthly_base_salary"]   = base_salary
            staff["position_allowance"]    = allowance_values["役職手当"]
            staff["attendance_allowance"]  = allowance_values["皆勤手当"]
            staff["housing_allowance"]     = allowance_values["住宅手当"]
            staff["commute_allowance"]     = allowance_values["交通費"]
            staff["social_insurance"]      = bool(social_var.get())
            staff["employment_insurance"]  = bool(employment_var.get())
            staff["birth_date"]            = birth
            staff["dependents"]            = dependents
            staff["has_spouse_deduction"]  = bool(spouse_var.get())
            staff["tax_table"]             = tax_table_var.get()
            staff["resident_tax_monthly"]  = resident
            _save_staff_list(self.staff_list)
            self.set_status(f"「{name}」の給与設定を保存しました。")
            if on_saved:
                on_saved()
            dlg.destroy()

        btn_f = tk.Frame(dlg, bg=C_BG)
        btn_f.pack(pady=(4, 10))
        _color_btn(btn_f, "保存", save, C_BTN_CHECK, width=10).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_f, "閉じる", dlg.destroy, C_BTN_REFR, width=10).pack(side=tk.LEFT, padx=4)

    # ── ユーザー管理 ──────────────────────────────
    def open_user_manager(self):
        """ユーザー管理画面を開く"""
        dlg = tk.Toplevel(self)
        dlg.title("ユーザー管理")
        dlg.geometry("350x450")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        tk.Label(dlg, text="ユーザー一覧", bg=C_BG, font=FONT_BOLD).pack(pady=(10, 4))

        # ユーザーリスト表示フレーム
        list_frame = tk.Frame(dlg, bg=C_BG)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10)

        lb = tk.Listbox(list_frame, font=FONT_MAIN, selectbackground=C_SELECT, height=12, relief=tk.FLAT, bd=1)
        sb_lb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=lb.yview)
        lb.configure(yscrollcommand=sb_lb.set)

        def refresh_user_lb():
            lb.delete(0, tk.END)
            users = _load_users()
            for username in sorted(users.keys()):
                role = users[username].get("role", "ユーザー")
                lb.insert(tk.END, f"{username} ({role})")

        refresh_user_lb()
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_lb.pack(side=tk.RIGHT, fill=tk.Y)

        # ユーザー名入力フレーム
        input_frame = tk.Frame(dlg, bg=C_BG)
        input_frame.pack(fill=tk.X, padx=10, pady=(8, 4))
        tk.Label(input_frame, text="ユーザー名:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        username_var = tk.StringVar()
        username_entry = ttk.Entry(input_frame, textvariable=username_var, width=18)
        username_entry.pack(side=tk.LEFT, padx=(4, 0))

        # パスワード入力フレーム
        pwd_frame = tk.Frame(dlg, bg=C_BG)
        pwd_frame.pack(fill=tk.X, padx=10, pady=(0, 4))
        tk.Label(pwd_frame, text="パスワード:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        password_var = tk.StringVar()
        password_entry = ttk.Entry(pwd_frame, textvariable=password_var, width=18, show="*")
        password_entry.pack(side=tk.LEFT, padx=(4, 0))

        # ロール選択フレーム
        role_frame = tk.Frame(dlg, bg=C_BG)
        role_frame.pack(fill=tk.X, padx=10, pady=(0, 8))
        tk.Label(role_frame, text="ロール:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        role_var = tk.StringVar(value="ユーザー")
        ttk.Combobox(role_frame, textvariable=role_var, values=("ユーザー", "管理者"),
                     width=15, state="readonly").pack(side=tk.LEFT, padx=(4, 0))

        def add_user():
            """新しいユーザーを追加"""
            username = username_var.get().strip()
            password = password_var.get()
            
            if not username or not password:
                messagebox.showwarning("入力エラー", "ユーザー名とパスワードを入力してください。", parent=dlg)
                return
            
            users = _load_users()
            if username in users:
                messagebox.showwarning("重複", f"「{username}」は既に存在します。", parent=dlg)
                return
            
            users[username] = {
                "password_hash": _hash_password(password),
                "role": role_var.get()
            }
            _save_users(users)
            refresh_user_lb()
            username_var.set("")
            password_var.set("")
            username_entry.focus()
            messagebox.showinfo("成功", f"ユーザー「{username}」を作成しました。", parent=dlg)

        def delete_user():
            """選択されたユーザーを削除"""
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("削除", "削除するユーザーを選択してください。", parent=dlg)
                return
            
            user_text = lb.get(sel[0])
            username = user_text.split(" (")[0]
            
            if username == "admin":
                messagebox.showerror("エラー", "管理者ユーザーは削除できません。", parent=dlg)
                return
            
            if messagebox.askyesno("確認", f"ユーザー「{username}」を削除しますか？", parent=dlg):
                users = _load_users()
                if username in users:
                    del users[username]
                    _save_users(users)
                    refresh_user_lb()
                    messagebox.showinfo("成功", f"ユーザー「{username}」を削除しました。", parent=dlg)

        def change_password():
            """選択されたユーザーのパスワードを変更"""
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("変更", "変更するユーザーを選択してください。", parent=dlg)
                return
            
            new_pwd = password_var.get()
            if not new_pwd:
                messagebox.showwarning("入力エラー", "新しいパスワードを入力してください。", parent=dlg)
                return
            
            user_text = lb.get(sel[0])
            username = user_text.split(" (")[0]
            
            if messagebox.askyesno("確認", f"ユーザー「{username}」のパスワードを変更しますか？", parent=dlg):
                users = _load_users()
                if username in users:
                    users[username]["password_hash"] = _hash_password(new_pwd)
                    _save_users(users)
                    password_var.set("")
                    messagebox.showinfo("成功", f"パスワードを変更しました。", parent=dlg)

        # ボタンフレーム
        btn_frame = tk.Frame(dlg, bg=C_BG)
        btn_frame.pack(pady=(6, 10))
        _color_btn(btn_frame, "作成", add_user, C_BTN_CHECK, width=8).pack(side=tk.LEFT, padx=2)
        _color_btn(btn_frame, "パスワード変更", change_password, C_BTN_SAVE, width=14).pack(side=tk.LEFT, padx=2)
        _color_btn(btn_frame, "削除", delete_user, C_BTN_DEL, width=8).pack(side=tk.LEFT, padx=2)

    def _clear_inputs(self):
        self.name_combo.set("")
        self.date_var.set(datetime.date.today().isoformat())
        self.checkin_time_var.set("")
        self.checkout_time_var.set("")
        self.break_var.set("0")
        self.admin_hours_var.set("0")
        self.admin_minutes_var.set("0")
        self.lessons_main_var.set("0")
        self.lessons_sub_var.set("0")
        self.actual_hours_var.set("-")
        self.total_work_var.set("-")
        self.night_hours_var.set("-")
        self.break_display_var.set("0分")
        self._edit_original_date = None

    # ── ヘルパー ─────────────────────────────────
    def parse_date(self):
        text = self.date_var.get().strip()
        if not text:
            return None
        try:
            return datetime.datetime.fromisoformat(text)
        except ValueError:
            try:
                d = datetime.date.fromisoformat(text)
                return datetime.datetime.combine(d, datetime.datetime.now().time())
            except ValueError:
                raise ValueError("日付フォーマットが不正です（YYYY-MM-DD）")

    def parse_names(self):
        text = self.name_combo.get().strip()
        if not text:
            raise ValueError("スタッフ名を選択してください。")
        return [text]

    def format_timestamp(self, ts):
        if not ts:
            return "-"
        try:
            return datetime.datetime.fromisoformat(ts).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return ts

    def format_time(self, ts):
        if not ts:
            return ""
        try:
            return datetime.datetime.fromisoformat(ts).strftime("%H:%M")
        except Exception:
            return ts

    def format_lessons(self, row):
        main = int(row.get("lessons_main") or 0)
        sub  = int(row.get("lessons_sub")  or 0)
        if main == 0 and sub == 0:
            return "0"
        return f"メイン{main} / サブ{sub}"

    def format_admin_time(self, row):
        minutes = int(row.get("admin_minutes") or 0)
        if minutes == 0:
            return "0分"
        return f"{minutes // 60}時間{minutes % 60}分"

    def format_break_time(self, row):
        m = int(row.get("break_minutes") or 0)
        return f"{m // 60}:{m % 60:02d}"

    def format_total_work(self, row):
        v = attendance.calculate_actual_hours(row)
        return f"{v:.2f}".rstrip("0").rstrip(".")

    def parse_timestamp(self, text):
        text = text.strip()
        if not text or text == "-":
            return None
        try:
            return datetime.datetime.fromisoformat(text)
        except ValueError:
            pass
        try:
            t = datetime.time.fromisoformat(text)
            d = self.parse_date() or datetime.datetime.now()
            return datetime.datetime.combine(d.date(), t)
        except ValueError:
            pass
        # ここまでの厳密な ISO 形式（時・分がゼロ埋めされている必要がある）で
        # 解釈できなかった場合、"2:00" や "2026-6-2 9:5" のようにゼロ埋め
        # なしで入力された時刻・日時も許容して解釈する。
        parsed = self._parse_flexible_datetime(text)
        if parsed is not None:
            return parsed
        raise ValueError("日時フォーマットが不正です（YYYY-MM-DD HH:MM）")

    def _parse_flexible_datetime(self, text):
        """ゼロ埋めなしの時刻・日時表記（例: "2:00", "2026-6-2 9:5"）を解釈する。

        解釈できない場合は None を返す（parse_timestamp 側でエラーメッセージを出す）。
        """
        date_part, time_part = None, text
        for sep in (" ", "T"):
            if sep in text:
                date_part, _, time_part = text.partition(sep)
                break

        try:
            time_bits = time_part.split(":")
            if not (2 <= len(time_bits) <= 3):
                return None
            hour, minute = int(time_bits[0]), int(time_bits[1])
            second = int(time_bits[2]) if len(time_bits) == 3 else 0
            t = datetime.time(hour=hour, minute=minute, second=second)
        except (ValueError, IndexError):
            return None

        if date_part:
            try:
                y, mo, da = (int(x) for x in date_part.split("-"))
                d = datetime.date(y, mo, da)
            except (ValueError, IndexError):
                return None
        else:
            base = self.parse_date() or datetime.datetime.now()
            d = base.date()

        return datetime.datetime.combine(d, t)

    def _apply_overnight_rollover(self, checkin_dt, checkout_dt):
        """退勤時刻が出勤時刻以前の場合、日をまたぐ勤務とみなして退勤日を1日進める。

        「出勤23:00・退勤2:00」のように時刻だけを入力した場合、
        parse_timestamp は両方とも同じ日付として解釈してしまうため、
        ここで退勤 <= 出勤 のケースを検出して補正する。
        """
        if checkin_dt and checkout_dt and checkout_dt <= checkin_dt:
            checkout_dt = checkout_dt + datetime.timedelta(days=1)
        return checkout_dt

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        vals = self.tree.item(selected[0], "values")
        if len(vals) < 5:
            return
        # cols: date, name, work_type, checkin, checkout, ...
        date, name, work_type, checkin, checkout = vals[0], vals[1], vals[2], vals[3], vals[4]

        self.name_combo.set(name)
        self.name_var.set(name)
        self._edit_original_date = date
        self.date_var.set(date)
        self.work_type_var.set(work_type or "通常出勤")

        self.checkin_time_var.set("" if checkin == "-" else checkin)
        self.checkout_time_var.set("" if checkout == "-" else checkout)

        rows = attendance.read_rows()
        row = attendance.find_row(rows, name, date)
        if row:
            bk = int(row.get("break_minutes") or 0)
            self.break_var.set(str(bk))
            self.break_display_var.set(f"{bk // 60}:{bk % 60:02d}")
            am = int(row.get("admin_minutes") or 0)
            self.admin_hours_var.set(str(am // 60))
            self.admin_minutes_var.set(str(am % 60))
            self.actual_hours_var.set(f"{attendance.calculate_actual_hours(row):.2f}h")
            self.total_work_var.set(self.format_total_work(row))
            self.night_hours_var.set(f"{attendance.calculate_night_hours(row):.2f}h")
            self.lessons_main_var.set(str(int(row.get("lessons_main") or 0)))
            self.lessons_sub_var.set(str(int(row.get("lessons_sub") or 0)))

    def update_time_labels(self):
        self.checkin_time_var.set("-")
        self.checkout_time_var.set("-")
        self.actual_hours_var.set("-")
        self.total_work_var.set("-")
        try:
            names = self.parse_names()
            if len(names) != 1:
                for v in (self.checkin_time_var, self.checkout_time_var,
                           self.actual_hours_var, self.total_work_var):
                    v.set("複数選択中")
                return
            when = self.parse_date()
            if not when:
                return
            row = attendance.find_row(attendance.read_rows(), names[0], when.date().isoformat())
            if not row:
                return
            self.checkin_time_var.set(self.format_timestamp(row.get("checkin")))
            self.checkout_time_var.set(self.format_timestamp(row.get("checkout")))
            self.break_var.set(str(int(row.get("break_minutes") or 0)))
            m = int(row.get("break_minutes") or 0)
            self.break_display_var.set(f"{m // 60}:{m % 60:02d}")
            am = int(row.get("admin_minutes") or 0)
            self.admin_hours_var.set(str(am // 60))
            self.admin_minutes_var.set(str(am % 60))
            self.actual_hours_var.set(f"{attendance.calculate_actual_hours(row):.2f}h")
            self.total_work_var.set(self.format_total_work(row))
            self.night_hours_var.set(f"{attendance.calculate_night_hours(row):.2f}h")
        except Exception:
            pass

    def set_status(self, message, error=False):
        self.status_var.set(message)
        self.status_label.configure(fg=C_STATUS_NG if error else C_STATUS_OK)

    # ── イベントハンドラ ──────────────────────────
    def on_register(self):
        """全項目を入力してから一括登録する。"""
        try:
            names = self.parse_names()
            if len(names) != 1:
                raise ValueError("スタッフを1名選択してください。")
            name = names[0]
            dt = self.parse_date()
            if not dt:
                raise ValueError("日付を入力してください。")
            date_str = dt.date().isoformat()
            work_type = self.work_type_var.get()

            checkin_dt  = self.parse_timestamp(self.checkin_time_var.get())
            checkout_dt = self.parse_timestamp(self.checkout_time_var.get())
            checkout_dt = self._apply_overnight_rollover(checkin_dt, checkout_dt)

            break_start_dt = None
            break_end_dt = None
            bk = int(self.break_var.get() or 0)
            ah        = int(self.admin_hours_var.get() or 0)
            am_val    = int(self.admin_minutes_var.get() or 0)
            admin_min = ah * 60 + am_val
            lessons_main = int(self.lessons_main_var.get() or 0)
            lessons_sub  = int(self.lessons_sub_var.get() or 0)

            am_obj  = labor.AttendanceManager()
            rows    = am_obj.read_rows()
            existing = labor.find_row(rows, name, date_str)

            if existing:
                if not messagebox.askyesno("上書き確認",
                        f"「{name}」{date_str} の記録が既に存在します。\n上書きしますか？"):
                    return
                existing["work_type"]     = work_type
                existing["checkin"]       = checkin_dt.isoformat()  if checkin_dt  else existing.get("checkin", "")
                existing["checkout"]      = checkout_dt.isoformat() if checkout_dt else ""
                existing["break_minutes"] = str(bk)
                existing["break_start"]   = break_start_dt.isoformat() if break_start_dt else existing.get("break_start", "")
                existing["break_end"]     = break_end_dt.isoformat()   if break_end_dt   else existing.get("break_end", "")
                existing["admin_minutes"] = str(admin_min)
                existing["lessons_main"]  = str(lessons_main)
                existing["lessons_sub"]   = str(lessons_sub)
            else:
                new_row = {
                    "id":            labor._next_id(rows),
                    "name":          name,
                    "date":          date_str,
                    "work_type":     work_type,
                    "checkin":       checkin_dt.isoformat()  if checkin_dt  else "",
                    "checkout":      checkout_dt.isoformat() if checkout_dt else "",
                    "break_minutes": str(bk),
                    "break_start":   break_start_dt.isoformat() if break_start_dt else "",
                    "break_end":     break_end_dt.isoformat()   if break_end_dt   else "",
                    "admin_minutes": str(admin_min),
                    "lessons_main":  str(lessons_main),
                    "lessons_sub":   str(lessons_sub),
                }
                rows.append(new_row)

            self._make_backup()
            am_obj.save_rows(rows)
            self.set_status(f"「{name}」{date_str} を登録しました。")
            self.refresh_records()
            self._clear_inputs()
            try:
                self._check_entry_labor_law(name, date_str)
            except Exception as exc:
                self.set_status(f"労基法チェックエラー: {exc}", error=True)
            try:
                self.check_overtime_alerts()
            except Exception as exc:
                self.set_status(f"36協定チェックエラー: {exc}", error=True)
        except ValueError as exc:
            messagebox.showerror("入力エラー", str(exc))
            self.set_status(str(exc), error=True)
        except Exception as exc:
            messagebox.showerror("エラー", str(exc))
            self.set_status(str(exc), error=True)

    def on_toggle_sort(self):
        self.sort_by_name = not self.sort_by_name
        self.sort_button.configure(text="日付順に戻す" if self.sort_by_name else "名前順ソート")
        self.refresh_records()

    def on_copy_last_pattern(self):
        """選択中スタッフの直近の勤務記録から、出勤・退勤時間と休憩時間をコピーする。"""
        name = self.name_combo.get().strip()
        if not name:
            messagebox.showwarning("前回コピー", "スタッフ名を選択してください。")
            return
        current_date = self.date_var.get().strip()
        rows = [
            r for r in attendance.read_rows()
            if r.get("name") == name and r.get("checkin") and r.get("date") != current_date
        ]
        if not rows:
            messagebox.showinfo("前回コピー", f"「{name}」の過去の勤務記録が見つかりません。")
            return
        rows.sort(key=lambda r: (r.get("date", ""), int(r.get("id") or 0)), reverse=True)
        last = rows[0]

        checkin_str  = self.format_time(last.get("checkin"))
        checkout_str = self.format_time(last.get("checkout"))
        self.checkin_time_var.set(checkin_str)
        self.checkout_time_var.set(checkout_str)

        bk = int(last.get("break_minutes") or 0)
        self.break_var.set(str(bk))
        self.break_display_var.set(f"{bk // 60}:{bk % 60:02d}")

        self.set_status(
            f"「{name}」{last.get('date')} の勤務時間（{checkin_str}〜{checkout_str}）をコピーしました。"
        )

    def on_add_lessons(self):
        try:
            names = self.parse_names()
            dt = self.parse_date()
            try:
                count_main = int(self.lessons_main_var.get())
                count_sub  = int(self.lessons_sub_var.get())
            except ValueError:
                err = "レッスン数は整数を入力してください。"
                messagebox.showerror("入力エラー", err)
                self.set_status(err, error=True)
                return
            if count_main == 0 and count_sub == 0:
                err = "メイン・サブいずれかのレッスン数を入力してください。"
                messagebox.showwarning("入力エラー", err)
                self.set_status(err, error=True)
                return
            msgs = []
            for name in names:
                when = datetime.datetime.combine(dt.date(), datetime.datetime.now().time()) if dt else None
                if count_main:
                    msgs.append(attendance.add_lessons(name, count_main, when=when, role="メイン"))
                if count_sub:
                    msgs.append(attendance.add_lessons(name, count_sub, when=when, role="サブ"))
            self.set_status("; ".join(msgs))
            self.refresh_records()
        except Exception as exc:
            messagebox.showerror("エラー", str(exc))
            self.set_status(str(exc), error=True)

    def refresh_records(self):
        all_rows = attendance.read_rows()

        # フィルター適用
        month_f = self.filter_month_var.get().strip()
        name_f  = self.filter_name_var.get().strip()
        rows = [r for r in all_rows
                if (not month_f or r.get("date", "").startswith(month_f))
                and (not name_f or r.get("name") == name_f)]

        if self.sort_by_name:
            rows = sorted(rows, key=lambda r: (r["name"].lower(), r["date"], int(r["id"])))
        else:
            rows = sorted(rows, key=lambda r: (r["date"], int(r["id"])), reverse=True)

        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, r in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", tk.END, tags=(tag,), values=(
                r["date"],
                r["name"],
                r.get("work_type") or "通常出勤",
                self.format_timestamp(r.get("checkin")),
                self.format_timestamp(r.get("checkout")),
                self.format_break_time(r),
                self.format_total_work(r),
                self.format_lessons(r),
                self.format_admin_time(r),
            ))

        # フィルターのドロップダウン選択肢を更新
        months = sorted({r["date"][:7] for r in all_rows if r.get("date", "") >= "2000"}, reverse=True)
        self.filter_month_combo["values"] = [""] + months
        self.filter_name_combo_filter["values"] = [""] + sorted(
            {r.get("name", "") for r in all_rows if r.get("name")})

        total = len(all_rows)
        shown = len(rows)
        msg = f"{shown} 件表示" if shown == total else f"{shown} 件表示（全 {total} 件中）"
        self.set_status(msg)

    def clear_filter(self):
        self.filter_month_var.set("")
        self.filter_name_var.set("")
        self.refresh_records()

    def _make_backup(self):
        """データ保存前に日付付きバックアップを作成する。"""
        src = labor.DEFAULT_CSV
        if not os.path.exists(src):
            return
        today = datetime.date.today().strftime("%Y%m%d")
        backup_dir = os.path.join(os.path.dirname(src), "backup")
        os.makedirs(backup_dir, exist_ok=True)
        dst = os.path.join(backup_dir, f"attendance_backup_{today}.csv")
        try:
            shutil.copy2(src, dst)
        except Exception as exc:
            self.set_status(f"バックアップ作成に失敗: {exc}", error=True)

    def on_save_corrections(self):
        try:
            names = self.parse_names()
            if len(names) != 1:
                raise ValueError("修正は1名のみ指定してください。")
            dt = self.parse_date()
            when_date = dt.date().isoformat() if dt else None
            if not when_date:
                raise ValueError("日付を入力してください。")

            lookup_date = self._edit_original_date if self._edit_original_date else when_date
            new_date = when_date if when_date != lookup_date else None

            checkin  = self.parse_timestamp(self.checkin_time_var.get())
            checkout = self.parse_timestamp(self.checkout_time_var.get())

            # 日付が変わった場合、出勤・退勤の日付部分も新しい日付に合わせる
            if new_date:
                if checkin:
                    checkin = datetime.datetime.combine(dt.date(), checkin.time())
                if checkout:
                    checkout = datetime.datetime.combine(dt.date(), checkout.time())

            # 退勤が出勤以前の時刻になっている場合、日をまたぐ勤務とみなして退勤日を1日進める
            checkout = self._apply_overnight_rollover(checkin, checkout)

            break_start_dt = None
            break_end_dt = None
            bk = int(self.break_var.get() or 0)
            ah = int(self.admin_hours_var.get())
            am = int(self.admin_minutes_var.get())
            self._make_backup()
            msg = attendance.update_record(
                names[0], lookup_date,
                checkin=checkin, checkout=checkout,
                break_minutes=bk, break_start=break_start_dt, break_end=break_end_dt,
                admin_minutes=ah * 60 + am,
                work_type=self.work_type_var.get(),
                new_date=new_date,
            )
            self.set_status(msg)
            self.refresh_records()
            self._clear_inputs()
            try:
                self._check_entry_labor_law(names[0], when_date)
            except Exception as exc:
                self.set_status(f"労基法チェックエラー: {exc}", error=True)
            try:
                self.check_overtime_alerts()
            except Exception as exc:
                self.set_status(f"36協定チェックエラー: {exc}", error=True)
        except (ValueError, Exception) as exc:
            messagebox.showerror("エラー", str(exc))
            self.set_status(str(exc), error=True)

    def on_delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("削除", "削除するレコードを選択してください。")
            return
        if not messagebox.askyesno("確認", "選択したレコードを削除しますか？"):
            return
        self._make_backup()
        msgs = []
        for item in selected:
            vals = self.tree.item(item, "values")
            if len(vals) >= 2:
                msgs.append(attendance.delete_record(vals[1], vals[0]))
        self.set_status("; ".join(msgs))
        self.refresh_records()

    def show_report(self):
        rows = attendance.read_rows()
        by_name = {}
        for r in rows:
            name = r["name"]
            if name not in by_name:
                by_name[name] = {"hours": 0.0, "lessons_main": 0, "lessons_sub": 0, "night_hours": 0.0}
            by_name[name]["hours"]        += attendance.calculate_actual_hours(r)
            by_name[name]["night_hours"]  += attendance.calculate_night_hours(r)
            by_name[name]["lessons_main"] += int(r.get("lessons_main") or 0)
            by_name[name]["lessons_sub"]  += int(r.get("lessons_sub")  or 0)

        lines = [
            f"{n}: 勤務 {v['hours']:.2f}h / 深夜 {v['night_hours']:.2f}h / レッスン メイン{v['lessons_main']} サブ{v['lessons_sub']}"
            for n, v in sorted(by_name.items())
        ]
        messagebox.showinfo("集計結果", "\n".join(lines) if lines else "記録がありません。")

    # ── 36協定管理ウィンドウ ─────────────────────
    def open_labor_manager(self):
        dlg = tk.Toplevel(self)
        dlg.title("36協定管理")
        dlg.geometry("760x420")
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        top = tk.Frame(dlg, bg=C_BG)
        top.pack(fill=tk.X, padx=8, pady=6)

        tk.Label(top, text="対象年月(YYYY-MM):", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        ym_var = tk.StringVar(value=datetime.date.today().strftime("%Y-%m"))
        ttk.Entry(top, textvariable=ym_var, width=12).pack(side=tk.LEFT, padx=(6, 8))
        def do_refresh():
            v = ym_var.get().strip()
            try:
                y, m = v.split("-")
                y = int(y); m = int(m)
            except Exception:
                messagebox.showerror("入力エラー", "年月は YYYY-MM の形式で入力してください。", parent=dlg)
                return
            try:
                lm = labor.LaborAgreementManager(labor.AttendanceManager())
                data = lm.compute_monthly_summary(y, m)
            except Exception as exc:
                messagebox.showerror("エラー", str(exc), parent=dlg)
                return
            for it in tree.get_children():
                tree.delete(it)
            if not data:
                messagebox.showinfo("36協定管理", f"{v} の勤怠データがありません。", parent=dlg)
                return

            violations = []
            warnings_list = []

            for i, r in enumerate(data):
                level = r["warning_level"]
                if level in ("赤", "濃赤"):
                    tag = "row_red"
                elif level == "オレンジ":
                    tag = "row_orange"
                elif level == "黄色":
                    tag = "row_yellow"
                else:
                    tag = "even" if i % 2 == 0 else "odd"
                tree.insert(
                    "", tk.END, tags=(tag,),
                    values=(
                        r["name"],
                        r["total_hours"],
                        r["overtime_hours"],
                        r["night_hours"],
                        r["break_violations"],
                        r["judgement"],
                        r["warning_level"],
                    ),
                )

                name = r["name"]
                ot = r["overtime_hours"]
                category = self._get_staff_category(name)

                if category in ("年少者", "妊産婦"):
                    if r.get("night_hours", 0.0) > 0:
                        violations.append(f"・{name}さん({category})：深夜労働が記録されています")
                    if ot > 0:
                        violations.append(f"・{name}さん({category})：時間外労働が記録されています")

                if ot >= 100:
                    violations.append(f"・{name}さん：時間外 {ot}h（月100時間超 ─ 絶対的上限超過）")
                elif ot >= 80:
                    warnings_list.append(f"・{name}さん：時間外 {ot}h（月80時間超 ─ 特別条項限度超過）")
                elif ot >= 45:
                    warnings_list.append(f"・{name}さん：時間外 {ot}h（月45時間超 ─ 原則上限超過）")

                if r.get("break_violations", 0) > 0:
                    violations.append(f"・{name}さん：{r['break_violations']}件の休憩時間違反があります")

            if violations or warnings_list:
                parts = []
                if violations:
                    parts.append("【法令違反・上限超過】\n" + "\n".join(violations))
                if warnings_list:
                    parts.append("【要注意・警告】\n" + "\n".join(warnings_list))
                messagebox.showwarning("⚠ 36協定 違反・警告", "\n\n".join(parts), parent=dlg)

        ttk.Button(top, text="表示", command=do_refresh).pack(side=tk.LEFT)
        def export_csv():
            v = ym_var.get().strip()
            try:
                y, m = v.split("-")
                y = int(y); m = int(m)
            except Exception:
                messagebox.showerror("入力エラー", "年月は YYYY-MM の形式で入力してください。", parent=dlg)
                return
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")], parent=dlg)
            if not path:
                return
            lm = labor.LaborAgreementManager(labor.AttendanceManager())
            lm.export_csv(path, y, m)
            messagebox.showinfo("出力完了", f"CSVを保存しました。\n{path}", parent=dlg)
        ttk.Button(top, text="CSV出力", command=export_csv).pack(side=tk.LEFT, padx=6)
        def export_annual_csv():
            y = datetime.date.today().year
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")], parent=dlg)
            if not path:
                return
            lm = labor.LaborAgreementManager(labor.AttendanceManager())
            lm.export_annual_csv(path, y)
            messagebox.showinfo("出力完了", f"CSVを保存しました。\n{path}", parent=dlg)
        ttk.Button(top, text="年間CSV出力", command=export_annual_csv).pack(side=tk.LEFT, padx=6)
        def export_multi_csv():
            ym = avg_entry.get().strip()
            try:
                y, m = ym.split("-")
                y = int(y); m = int(m)
            except Exception:
                messagebox.showerror("入力エラー", "年月は YYYY-MM 形式で入力してください。", parent=dlg)
                return
            try:
                w = int(months_var.get() or 3)
            except Exception:
                messagebox.showerror("入力エラー", "月数は整数で入力してください。", parent=dlg)
                return
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")], parent=dlg)
            if not path:
                return
            try:
                lm = labor.LaborAgreementManager(labor.AttendanceManager())
                data = lm.compute_multi_month_average(y, m, window=w)
                if not data:
                    messagebox.showinfo("複数月CSV出力", "対象期間にデータがありません。", parent=dlg)
                    return
                lm.export_multi_month_csv(path, y, m, window=w)
                messagebox.showinfo("出力完了", f"CSVを保存しました。\n{path}", parent=dlg)
            except Exception as exc:
                messagebox.showerror("エラー", str(exc), parent=dlg)
        ttk.Button(top, text="複数月CSV出力", command=export_multi_csv).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="年間720チェック", command=lambda: self._do_annual_check(dlg)).pack(side=tk.LEFT, padx=6)
        tk.Label(top, text=" / 複数月平均(末月 YYYY-MM):", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT, padx=(12,4))
        avg_entry = ttk.Entry(top, width=8)
        avg_entry.pack(side=tk.LEFT)
        avg_entry.insert(0, datetime.date.today().strftime("%Y-%m"))
        months_var = tk.StringVar(value="3")
        ttk.Entry(top, textvariable=months_var, width=3).pack(side=tk.LEFT, padx=(6,0))
        ttk.Button(top, text="複数月平均チェック", command=lambda: self._do_multi_month_check(dlg, avg_entry.get(), int(months_var.get() or 3))).pack(side=tk.LEFT, padx=6)

        cols = ("name", "total", "overtime", "night", "breaks", "judgement", "level")
        tree = ttk.Treeview(dlg, columns=cols, show="headings", height=16)
        headers = {
            "name": ("スタッフ名", 180),
            "total": ("月間勤務時間", 110),
            "overtime": ("時間外労働時間", 110),
            "night": ("深夜時間", 100),
            "breaks": ("休憩違反件数", 100),
            "judgement": ("36協定判定", 120),
            "level": ("警告レベル", 110),
        }
        for cid, (htext, w) in headers.items():
            tree.heading(cid, text=htext)
            tree.column(cid, width=w)
        tree.tag_configure("odd",        background=C_ROW_ODD)
        tree.tag_configure("even",       background=C_ROW_EVEN)
        tree.tag_configure("row_yellow", background="#FFFACD")
        tree.tag_configure("row_orange", background="#FFE0B3")
        tree.tag_configure("row_red",    background="#FFB3B3")

        sb = ttk.Scrollbar(dlg, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8,0), pady=8)
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=8)

        # keep reference so other methods can update this Treeview
        self.labor_tree = tree

        do_refresh()

    def _do_annual_check(self, parent):
        y = datetime.date.today().year
        lm = labor.LaborAgreementManager(labor.AttendanceManager())
        data = lm.compute_annual_summary(y)
        over_720 = [f"{r['name']}: {r['total_hours']}h" for r in data if r['over_720']]
        over_360 = [f"{r['name']}: {r['annual_overtime_hours']}h" for r in data if r['over_360_overtime']]
        messages = []
        if over_720:
            messages.append("[年間720時間超過]\n" + "\n".join(over_720))
        if over_360:
            messages.append("[年間時間外360時間超過]\n" + "\n".join(over_360))
        if messages:
            messagebox.showwarning("年間集計警告", "\n\n".join(messages), parent=parent)
        else:
            messagebox.showinfo("年間720時間チェック", "超過者はいません。", parent=parent)

    def _do_multi_month_check(self, parent, ym_text, window):
        try:
            y, m = ym_text.split("-")
            y = int(y); m = int(m)
        except Exception:
            messagebox.showerror("入力エラー", "年月は YYYY-MM 形式で入力してください。", parent=parent)
            return
        lm = labor.LaborAgreementManager(labor.AttendanceManager())
        data = lm.compute_multi_month_average(y, m, window=window)
        tree = getattr(self, 'labor_tree', None)
        if tree is None:
            messagebox.showerror("内部エラー", "Treeview が見つかりません。36協定管理画面を再表示してください。", parent=parent)
            return
        # reconfigure tree to show multi-month average
        tree_cols = ("name", "avg", "window", "flag")
        tree.config(columns=tree_cols)
        tree.delete(*tree.get_children())
        tree.heading("name", text="スタッフ名")
        tree.column("name", width=260)
        tree.heading("avg", text=f"平均時間外(h)")
        tree.column("avg", width=140)
        tree.heading("window", text="ウィンドウ(月)")
        tree.column("window", width=100)
        tree.heading("flag", text="超過フラグ")
        tree.column("flag", width=120)
        over_list = []
        for i, r in enumerate(data):
            is_over = r.get("over_80_avg", False)
            tag = "row_red" if is_over else ("even" if i % 2 == 0 else "odd")
            flag = "超過" if is_over else "正常"
            tree.insert("", tk.END, tags=(tag,), values=(r["name"], r["average_overtime"], r.get("window", window), flag))
            if is_over:
                over_list.append(f"・{r['name']}さん：平均時間外 {r['average_overtime']}h（80時間超）")

        if over_list:
            messagebox.showwarning(
                "⚠ 複数月平均 時間外警告",
                "【月平均80時間超過 ─ 特別条項の限度超過】\n\n" + "\n".join(over_list),
                parent=parent,
            )
        else:
            messagebox.showinfo("複数月平均チェック", "超過者はいません。", parent=parent)

    def check_overtime_alerts(self):
        today = datetime.date.today()
        lm = labor.LaborAgreementManager(labor.AttendanceManager())
        monthly = lm.compute_monthly_summary(today.year, today.month)
        annual = lm.compute_annual_summary(today.year)

        violations = []  # 法令違反・絶対的上限超過
        warnings = []    # 要注意・危険域

        for r in monthly:
            name = r["name"]
            ot = r["overtime_hours"]
            category = self._get_staff_category(name)

            # 年少者・妊産婦の法令違反チェック
            if category in ("年少者", "妊産婦"):
                if r.get("night_hours", 0.0) > 0:
                    violations.append(f"・{name}さん({category})：深夜労働が記録されています")
                if ot > 0:
                    violations.append(f"・{name}さん({category})：時間外労働が記録されています")

            # 月間時間外労働の段階別チェック
            if ot >= 100:
                violations.append(f"・{name}さん：今月の時間外労働 {ot}h（月100時間超 ─ 絶対的上限超過）")
            elif ot >= 80:
                warnings.append(f"・{name}さん：今月の時間外労働 {ot}h（月80時間超 ─ 特別条項限度超過）")
            elif ot >= 45:
                warnings.append(f"・{name}さん：今月の時間外労働 {ot}h（月45時間超 ─ 原則上限超過）")

            # 休憩時間違反チェック
            bv = r.get("break_violations", 0)
            if bv > 0:
                violations.append(f"・{name}さん：今月 {bv}件の休憩時間違反があります（労基法34条）")

        # 年間チェック
        for r in annual:
            name = r["name"]
            if r.get("over_360_overtime"):
                violations.append(
                    f"・{name}さん：年間時間外労働 {r['annual_overtime_hours']}h（360時間超 ─ 年間上限超過）"
                )
            if r.get("over_720"):
                warnings.append(
                    f"・{name}さん：年間総労働時間 {r['total_hours']}h（720時間超過）"
                )

        if (violations or warnings) and not self._overtime_alert_shown:
            self._overtime_alert_shown = True
            parts = []
            if violations:
                parts.append("【法令違反・上限超過】\n" + "\n".join(violations))
            if warnings:
                parts.append("【要注意・警告】\n" + "\n".join(warnings))
            messagebox.showwarning("⚠ 36協定 違反・警告", "\n\n".join(parts))

    # ── 労働基準法チェック（1日8h・週40h・休憩） ──────
    def _check_entry_labor_law(self, name, date_str):
        """登録・保存した1件について労働基準法の法定労働時間・休憩を確認する。"""
        am_obj = labor.AttendanceManager()
        rows = am_obj.read_rows()
        row = labor.find_row(rows, name, date_str)
        if not row:
            return

        violations = []

        # 1日の法定労働時間チェック（実労働時間 > 8h）
        actual_h = round(labor.calculate_actual_hours(row), 2)
        if actual_h > 8.0:
            over = round(actual_h - 8.0, 2)
            violations.append(
                f"・{name}さん（{date_str}）：実労働時間 {actual_h}h"
                f"（法定上限1日8時間超 ／ {over}h超過）"
            )

        # 休憩時間チェック（労基法34条）
        # 判定は総勤務時間（出退勤の差）で行う
        total_h = am_obj.calculate_total_hours(row)
        break_min = int(row.get("break_minutes") or 0)
        if total_h > 8.0 and break_min < 60:
            violations.append(
                f"・{name}さん（{date_str}）：8時間超勤務の休憩 {break_min}分"
                f"（60分以上必要 ／ {60 - break_min}分不足）"
            )
        elif total_h > 6.0 and break_min < 45:
            violations.append(
                f"・{name}さん（{date_str}）：6時間超勤務の休憩 {break_min}分"
                f"（45分以上必要 ／ {45 - break_min}分不足）"
            )

        # 週40時間チェック
        try:
            date_obj = datetime.date.fromisoformat(date_str)
            cal = date_obj.isocalendar()
            week_key = (cal[0], cal[1])
            week_total = 0.0
            for r in rows:
                try:
                    d = datetime.date.fromisoformat(r.get("date", ""))
                    dc = d.isocalendar()
                    if r.get("name") == name and (dc[0], dc[1]) == week_key:
                        week_total += labor.calculate_actual_hours(r)
                except Exception:
                    continue
            week_total = round(week_total, 2)
            if week_total > 40.0:
                over_w = round(week_total - 40.0, 2)
                violations.append(
                    f"・{name}さん（{date_str}の週）：週の実労働時間 {week_total}h"
                    f"（法定上限週40時間超 ／ {over_w}h超過）"
                )
        except Exception:
            pass

        if violations:
            messagebox.showwarning(
                "⚠ 労働基準法 違反・警告",
                "以下の法定労働時間・休憩規定違反が検出されました。\n\n" + "\n".join(violations),
            )

    # ── Excel 出力 ────────────────────────────────
    def _ask_export_month(self, all_rows):
        """月選択ダイアログを表示し、選択された YYYY-MM 文字列（または None でキャンセル）を返す。"""
        months = sorted(
            {r["date"][:7] for r in all_rows if r.get("date", "") >= "2000"},
            reverse=True,
        )
        if not months:
            return None

        result = {"value": None}
        dlg = tk.Toplevel(self)
        dlg.title("出力月を選択")
        dlg.geometry("280x140")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        tk.Label(dlg, text="出力対象の年月を選択してください", bg=C_BG, font=FONT_MAIN).pack(pady=(14, 6))

        var = tk.StringVar(value=months[0])
        combo = ttk.Combobox(dlg, textvariable=var, values=months, state="readonly", width=14)
        combo.pack()

        def on_ok():
            result["value"] = var.get()
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btn_frame = tk.Frame(dlg, bg=C_BG)
        btn_frame.pack(pady=14)
        _color_btn(btn_frame, "OK",       on_ok,     C_BTN_CHECK, width=8).pack(side=tk.LEFT, padx=6)
        _color_btn(btn_frame, "キャンセル", on_cancel, C_BTN_DEL,   width=10).pack(side=tk.LEFT, padx=6)

        dlg.bind("<Return>", lambda _: on_ok())
        dlg.bind("<Escape>", lambda _: on_cancel())
        self.wait_window(dlg)
        return result["value"]

    def export_excel(self):
        all_rows = attendance.read_rows()
        if not all_rows:
            messagebox.showwarning("Excel出力", "出力するデータがありません。")
            return

        month_str = self._ask_export_month(all_rows)
        if month_str is None:
            return

        rows = [r for r in all_rows if r.get("date", "").startswith(month_str)]
        if not rows:
            messagebox.showwarning("Excel出力", f"{month_str} のデータがありません。")
            return

        default_name = f"勤怠集計_{month_str}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Excel ファイルの保存先",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx")],
        )
        if not path:
            return

        wb = openpyxl.Workbook()

        # ── シート1: 詳細レコード ─────────────────
        ws1 = wb.active
        ws1.title = "勤怠詳細"

        hdr_fill  = PatternFill("solid", fgColor="4A7FD4")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Meiryo UI", size=10)
        body_font = Font(name="Meiryo UI", size=10)
        even_fill = PatternFill("solid", fgColor="EAF0FB")
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center")
        thin      = Side(style="thin", color="CCCCCC")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        detail_headers = ["日付", "スタッフ名", "勤務区分", "出勤時間", "退勤時間",
                          "休憩(分)", "総勤務(h)", "メインレッスン", "サブレッスン", "事務作業(分)"]
        col_widths     = [14, 16, 14, 22, 22, 10, 12, 14, 14, 14]

        for ci, (h, w) in enumerate(zip(detail_headers, col_widths), 1):
            cell = ws1.cell(row=1, column=ci, value=h)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = border
            ws1.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        sorted_rows = sorted(rows, key=lambda r: (r["date"], int(r["id"])), reverse=True)
        for ri, r in enumerate(sorted_rows, 2):
            fill = even_fill if ri % 2 == 0 else PatternFill()
            vals = [
                r.get("date", ""),
                r.get("name", ""),
                r.get("work_type") or "通常出勤",
                self.format_timestamp(r.get("checkin")),
                self.format_timestamp(r.get("checkout")),
                int(r.get("break_minutes") or 0),
                round(attendance.calculate_actual_hours(r), 2),
                int(r.get("lessons_main") or 0),
                int(r.get("lessons_sub")  or 0),
                int(r.get("admin_minutes") or 0),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws1.cell(row=ri, column=ci, value=v)
                cell.font      = body_font
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center if ci != 2 else left
        ws1.freeze_panes = "A2"

        # ── シート2: スタッフ別集計（勤務区分別） ────────────────
        ws2 = wb.create_sheet("スタッフ別集計")

        fill_normal    = PatternFill("solid", fgColor="4A7FD4")
        fill_specified = PatternFill("solid", fgColor="C87D00")
        fill_legal     = PatternFill("solid", fgColor="A93226")
        total_fill     = PatternFill("solid", fgColor="D5E8D4")
        total_font     = Font(bold=True, name="Meiryo UI", size=10)

        # ─ 2行ヘッダー
        # 列レイアウト:
        # 1=スタッフ名, 2-3=通常出勤, 4-5=所定休日出勤, 6-7=法定休日出勤,
        # 8=総勤務時間, 9=メインL, 10=サブL, 11=総L, 12=事務作業

        def _ws2_hdr(col, text, row1_fill, merge_end_col=None, row2_sub=None):
            if merge_end_col:
                ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=merge_end_col)
            else:
                ws2.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            c = ws2.cell(row=1, column=col, value=text)
            c.fill = row1_fill; c.font = hdr_font; c.alignment = center; c.border = border
            if row2_sub:
                for offset, sub in enumerate(row2_sub):
                    c2 = ws2.cell(row=2, column=col + offset, value=sub)
                    c2.fill = row1_fill; c2.font = hdr_font; c2.alignment = center; c2.border = border

        _ws2_hdr(1,  "スタッフ名",   hdr_fill)
        _ws2_hdr(2,  "通常出勤",     fill_normal,    merge_end_col=3,  row2_sub=["日数", "時間(h)"])
        _ws2_hdr(4,  "所定休日出勤", fill_specified, merge_end_col=5,  row2_sub=["日数", "時間(h)"])
        _ws2_hdr(6,  "法定休日出勤", fill_legal,     merge_end_col=7,  row2_sub=["日数", "時間(h)"])
        _ws2_hdr(8,  "総勤務時間(h)",  hdr_fill)
        _ws2_hdr(9,  "メインレッスン", hdr_fill)
        _ws2_hdr(10, "サブレッスン",   hdr_fill)
        _ws2_hdr(11, "総レッスン",     hdr_fill)
        _ws2_hdr(12, "事務作業(h)",    hdr_fill)

        col_widths_s2 = [18, 8, 10, 10, 12, 10, 12, 14, 14, 12, 10, 12]
        for ci, w in enumerate(col_widths_s2, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        # ─ データ集計
        by_name = {}
        for r in rows:
            n = r["name"]
            if n not in by_name:
                by_name[n] = {
                    "days_n": 0, "hours_n": 0.0,
                    "days_s": 0, "hours_s": 0.0,
                    "days_l": 0, "hours_l": 0.0,
                    "hours_total": 0.0,
                    "main": 0, "sub": 0, "admin": 0,
                }
            h = attendance.calculate_actual_hours(r)
            by_name[n]["hours_total"] += h
            by_name[n]["main"]  += int(r.get("lessons_main")  or 0)
            by_name[n]["sub"]   += int(r.get("lessons_sub")   or 0)
            by_name[n]["admin"] += int(r.get("admin_minutes") or 0)
            wt = r.get("work_type") or "通常出勤"
            if wt == "所定休日出勤":
                by_name[n]["days_s"] += 1; by_name[n]["hours_s"] += h
            elif wt == "法定休日出勤":
                by_name[n]["days_l"] += 1; by_name[n]["hours_l"] += h
            else:
                by_name[n]["days_n"] += 1; by_name[n]["hours_n"] += h

        # ─ データ行（3行目〜）
        for ri, (name, v) in enumerate(sorted(by_name.items()), 3):
            fill = even_fill if ri % 2 == 0 else PatternFill()
            row_vals = [
                name,
                v["days_n"],  round(v["hours_n"], 2),
                v["days_s"],  round(v["hours_s"], 2),
                v["days_l"],  round(v["hours_l"], 2),
                round(v["hours_total"], 2),
                v["main"], v["sub"], v["main"] + v["sub"],
                round(v["admin"] / 60, 2),
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.font = body_font; c.fill = fill; c.border = border
                c.alignment = left if ci == 1 else center
        ws2.freeze_panes = "A3"

        # ─ 合計行
        total_row = len(by_name) + 3
        total_vals = [
            "【合計】",
            sum(v["days_n"]  for v in by_name.values()),
            round(sum(v["hours_n"]  for v in by_name.values()), 2),
            sum(v["days_s"]  for v in by_name.values()),
            round(sum(v["hours_s"]  for v in by_name.values()), 2),
            sum(v["days_l"]  for v in by_name.values()),
            round(sum(v["hours_l"]  for v in by_name.values()), 2),
            round(sum(v["hours_total"] for v in by_name.values()), 2),
            sum(v["main"] for v in by_name.values()),
            sum(v["sub"]  for v in by_name.values()),
            sum(v["main"] + v["sub"] for v in by_name.values()),
            round(sum(v["admin"] for v in by_name.values()) / 60, 2),
        ]
        for ci, val in enumerate(total_vals, 1):
            c = ws2.cell(row=total_row, column=ci, value=val)
            c.font = total_font; c.fill = total_fill; c.border = border
            c.alignment = left if ci == 1 else center

        # ── シート3: 労働基準法違反レポート ─────────────────
        ws3 = wb.create_sheet("労基法違反レポート")

        hdr3_fill       = PatternFill("solid", fgColor="D35400")
        hour_viol_fill  = PatternFill("solid", fgColor="FFCCCC")
        break_viol_fill = PatternFill("solid", fgColor="FFF2CC")
        ok_fill         = PatternFill("solid", fgColor="D5F5E3")

        headers3    = ["スタッフ名", "日付 / 対象週", "違反種別", "実績値", "法定基準", "超過 / 不足"]
        col_widths3 = [16, 18, 30, 14, 14, 14]
        for ci, (h, w) in enumerate(zip(headers3, col_widths3), 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = hdr3_fill; c.font = hdr_font; c.alignment = center; c.border = border
            ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        am3 = labor.AttendanceManager()
        violation_entries = []

        # ─ 1日ごとのチェック（1日8h超・休憩不足）
        for r in sorted_rows:
            r_name   = r.get("name", "")
            r_date   = r.get("date", "")
            actual_h = round(labor.calculate_actual_hours(r), 2)
            total_h  = am3.calculate_total_hours(r)
            bk_min   = int(r.get("break_minutes") or 0)

            if actual_h > 8.0:
                over = round(actual_h - 8.0, 2)
                violation_entries.append((
                    r_name, r_date,
                    "1日8時間超（法定労働時間）",
                    f"{actual_h}h", "8.0h以内", f"+{over}h超過", "hour",
                ))

            if total_h > 8.0 and bk_min < 60:
                violation_entries.append((
                    r_name, r_date,
                    "休憩不足（8時間超 → 60分必要）",
                    f"{bk_min}分", "60分以上", f"{60 - bk_min}分不足", "break",
                ))
            elif total_h > 6.0 and bk_min < 45:
                violation_entries.append((
                    r_name, r_date,
                    "休憩不足（6時間超 → 45分必要）",
                    f"{bk_min}分", "45分以上", f"{45 - bk_min}分不足", "break",
                ))

        # ─ 週ごとのチェック（週40時間超）
        checked_weeks = set()
        for r in sorted_rows:
            r_name = r.get("name", "")
            try:
                d = datetime.date.fromisoformat(r.get("date", ""))
                cal = d.isocalendar()
                wk_key = (r_name, cal[0], cal[1])
                if wk_key in checked_weeks:
                    continue
                checked_weeks.add(wk_key)
                week_key = (cal[0], cal[1])
                wk_total = 0.0
                for rw in all_rows:
                    if rw.get("name") != r_name:
                        continue
                    try:
                        d2 = datetime.date.fromisoformat(rw.get("date", ""))
                        c2 = d2.isocalendar()
                        if (c2[0], c2[1]) == week_key:
                            wk_total += labor.calculate_actual_hours(rw)
                    except Exception:
                        continue
                wk_total = round(wk_total, 2)
                if wk_total > 40.0:
                    over = round(wk_total - 40.0, 2)
                    violation_entries.append((
                        r_name, f"{r.get('date', '')}の週",
                        "週40時間超（法定労働時間）",
                        f"{wk_total}h", "40.0h以内", f"+{over}h超過", "hour",
                    ))
            except Exception:
                continue

        # ─ 違反データを書き込む
        if violation_entries:
            for ri, (vname, vdate, vtype, vactual, vstd, vdiff, vlevel) in enumerate(violation_entries, 2):
                row_fill = hour_viol_fill if vlevel == "hour" else break_viol_fill
                for ci, val in enumerate([vname, vdate, vtype, vactual, vstd, vdiff], 1):
                    c = ws3.cell(row=ri, column=ci, value=val)
                    c.font = body_font; c.fill = row_fill; c.border = border
                    c.alignment = left if ci <= 2 else center
        else:
            c = ws3.cell(row=2, column=1, value="違反・警告事項はありません")
            c.font = Font(bold=True, name="Meiryo UI", size=10, color="1A7A3C")
            c.fill = ok_fill
            c.alignment = left

        # ─ 凡例
        legend_row = len(violation_entries) + 4 if violation_entries else 4
        ws3.cell(row=legend_row, column=1, value="【凡例】").font = Font(
            bold=True, name="Meiryo UI", size=9
        )
        leg1 = ws3.cell(row=legend_row + 1, column=1, value="赤背景：1日・週の実労働時間が法定上限超過")
        leg1.fill = hour_viol_fill
        leg1.font = Font(name="Meiryo UI", size=9)
        leg2 = ws3.cell(row=legend_row + 2, column=1, value="黄背景：休憩時間不足（労基法34条）")
        leg2.fill = break_viol_fill
        leg2.font = Font(name="Meiryo UI", size=9)

        ws3.freeze_panes = "A2"

        try:
            wb.save(path)
            self.set_status(f"Excel出力完了: {os.path.basename(path)}")
            if messagebox.askyesno("完了", f"保存しました。\n{path}\n\nファイルを開きますか？"):
                os.startfile(path)
        except PermissionError:
            messagebox.showerror("書き込みエラー", "ファイルが他のアプリで開かれています。閉じてから再試行してください。")
        except Exception as exc:
            messagebox.showerror("エラー", str(exc))

    # ── 有給休暇管理 ──────────────────────────────
    def open_paid_leave_manager(self):
        dlg = tk.Toplevel(self)
        dlg.title("有給休暇管理")
        dlg.geometry("800x540")
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        notebook = ttk.Notebook(dlg)
        notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        tab_balance = tk.Frame(notebook, bg=C_BG)
        tab_request = tk.Frame(notebook, bg=C_BG)
        tab_grant   = tk.Frame(notebook, bg=C_BG)
        tab_usage   = tk.Frame(notebook, bg=C_BG)
        tab_history = tk.Frame(notebook, bg=C_BG)
        notebook.add(tab_balance, text="残高一覧")
        notebook.add(tab_request, text="有給申請")
        notebook.add(tab_grant,   text="付与登録")
        notebook.add(tab_usage,   text="取得登録")
        notebook.add(tab_history, text="履歴一覧")

        C_GRANT = "#16A085"

        # ── タブ1: 残高一覧 ──────────────────────
        ctrl_bal = tk.Frame(tab_balance, bg=C_BG)
        ctrl_bal.pack(fill=tk.X, padx=8, pady=6)
        tk.Label(ctrl_bal, text="基準日:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        bal_date_var = tk.StringVar(value=datetime.date.today().isoformat())
        ttk.Entry(ctrl_bal, textvariable=bal_date_var, width=14).pack(side=tk.LEFT, padx=(4, 8))

        bal_cols = ("name", "granted", "used", "remaining", "alert")
        bal_tree = ttk.Treeview(tab_balance, columns=bal_cols, show="headings", height=14)
        for cid, ht, w, anch in [
            ("name",      "スタッフ名",      160, tk.W),
            ("granted",   "累計付与日数",    110, tk.CENTER),
            ("used",      "取得済日数",      110, tk.CENTER),
            ("remaining", "残日数",           90, tk.CENTER),
            ("alert",     "年5日義務",        140, tk.CENTER),
        ]:
            bal_tree.heading(cid, text=ht)
            bal_tree.column(cid, width=w, anchor=anch)
        bal_tree.tag_configure("odd",  background=C_ROW_ODD)
        bal_tree.tag_configure("even", background=C_ROW_EVEN)
        bal_tree.tag_configure("warn", background="#FFF3CD")

        sb_bal = ttk.Scrollbar(tab_balance, orient=tk.VERTICAL, command=bal_tree.yview)
        bal_tree.configure(yscroll=sb_bal.set)
        bal_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        sb_bal.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 8))

        def refresh_balance():
            as_of = bal_date_var.get().strip() or datetime.date.today().isoformat()
            year  = as_of[:4]
            data  = _load_paid_leave()
            for it in bal_tree.get_children():
                bal_tree.delete(it)
            staff_names = sorted([s.get("name", "") for s in self.staff_list if isinstance(s, dict)])
            for i, name in enumerate(staff_names):
                remaining, granted, used = _calc_paid_leave_balance(name, data, as_of)
                used_yr = sum(
                    float(u.get("days", 0)) for u in data["usages"]
                    if u["name"] == name and u.get("date", "").startswith(year)
                )
                need = round(5 - used_yr, 1)
                alert = "✓ 達成" if used_yr >= 5 else f"要取得(あと{need}日)"
                tag   = "warn" if used_yr < 5 else ("even" if i % 2 == 0 else "odd")
                bal_tree.insert("", tk.END, tags=(tag,), values=(name, granted, used, remaining, alert))

        _color_btn(ctrl_bal, "更新", refresh_balance, C_BTN_REFR, width=8).pack(side=tk.LEFT)
        refresh_balance()

        # ── タブ(申請): 有給申請 ──────────────────
        is_admin = _get_user_role(self.logged_in_user) == "管理者"

        rq = tk.Frame(tab_request, bg=C_BG)
        rq.pack(fill=tk.X, padx=12, pady=10)

        def rqlbl(text, row, col):
            tk.Label(rq, text=text, bg=C_BG, font=FONT_MAIN).grid(
                row=row, column=col, sticky=tk.W, padx=(0, 4), pady=4)

        rqlbl("スタッフ名:", 0, 0)
        rq_name_var = tk.StringVar()
        ttk.Combobox(rq, textvariable=rq_name_var,
                     values=[s.get("name") for s in self.staff_list if isinstance(s, dict)],
                     width=16, state="readonly").grid(row=0, column=1, sticky=tk.W)
        rqlbl("取得予定日:", 0, 2)
        rq_date_var = tk.StringVar(value=(datetime.date.today() + datetime.timedelta(days=1)).isoformat())
        ttk.Entry(rq, textvariable=rq_date_var, width=14).grid(row=0, column=3, sticky=tk.W)
        rqlbl("日数:", 1, 0)
        rq_days_var = tk.StringVar(value="1")
        ttk.Entry(rq, textvariable=rq_days_var, width=8).grid(row=1, column=1, sticky=tk.W)
        tk.Label(rq, text="(0.5=半日)", bg=C_BG, font=("Meiryo UI", 9)).grid(row=1, column=2, sticky=tk.W)
        rqlbl("理由・備考:", 2, 0)
        rq_note_var = tk.StringVar()
        ttk.Entry(rq, textvariable=rq_note_var, width=36).grid(
            row=2, column=1, columnspan=3, sticky=tk.W + tk.E)
        rq_remain_var = tk.StringVar(value="残日数: - 日")
        tk.Label(rq, textvariable=rq_remain_var, bg=C_BG, font=FONT_BOLD,
                 fg=C_GRANT).grid(row=0, column=4, padx=(16, 0), sticky=tk.W)

        def update_rq_remain_lbl(*_):
            name = rq_name_var.get().strip()
            if name:
                rem, _, _ = _calc_paid_leave_balance(name, _load_paid_leave())
                rq_remain_var.set(f"残日数: {rem} 日")
            else:
                rq_remain_var.set("残日数: - 日")
        rq_name_var.trace_add("write", update_rq_remain_lbl)

        request_cols = ("id", "name", "leave_date", "days", "note", "status", "requested_at", "decided_by")
        request_tree = ttk.Treeview(tab_request, columns=request_cols, show="headings", height=10)
        for cid, ht, w, anch in [
            ("id",           "ID",         45,  tk.CENTER),
            ("name",         "スタッフ名",  110, tk.W),
            ("leave_date",   "取得予定日",  100, tk.CENTER),
            ("days",         "日数",        60,  tk.CENTER),
            ("note",         "理由・備考",  160, tk.W),
            ("status",       "状態",        70,  tk.CENTER),
            ("requested_at", "申請日",      100, tk.CENTER),
            ("decided_by",   "承認者",       90,  tk.CENTER),
        ]:
            request_tree.heading(cid, text=ht)
            request_tree.column(cid, width=w, anchor=anch)
        request_tree.tag_configure("pending",  background="#FFF3CD")
        request_tree.tag_configure("approved", background="#D5F5E3")
        request_tree.tag_configure("rejected", background="#FADBD8")
        sb_request = ttk.Scrollbar(tab_request, orient=tk.VERTICAL, command=request_tree.yview)
        request_tree.configure(yscroll=sb_request.set)

        def refresh_request_tree():
            data = _load_paid_leave()
            for it in request_tree.get_children():
                request_tree.delete(it)
            for r in sorted(data["requests"], key=lambda x: x.get("requested_at", ""), reverse=True):
                status = r.get("status", "申請中")
                tag = "pending" if status == "申請中" else ("approved" if status == "承認" else "rejected")
                request_tree.insert("", tk.END, tags=(tag,), values=(
                    r.get("id", ""), r.get("name", ""), r.get("leave_date", ""),
                    r.get("days", ""), r.get("note", ""), status,
                    r.get("requested_at", ""), r.get("decided_by", "")))

        def submit_request():
            name = rq_name_var.get().strip()
            if not name:
                messagebox.showwarning("入力エラー", "スタッフ名を選択してください。", parent=dlg)
                return
            try:
                datetime.date.fromisoformat(rq_date_var.get().strip())
                days = float(rq_days_var.get().strip())
                if days <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "取得予定日と日数（正の数）を正しく入力してください。", parent=dlg)
                return
            data = _load_paid_leave()
            data["requests"].append({
                "id":           _next_pl_id(data["requests"]),
                "name":         name,
                "leave_date":   rq_date_var.get().strip(),
                "days":         days,
                "note":         rq_note_var.get().strip(),
                "status":       "申請中",
                "requested_at": datetime.date.today().isoformat(),
                "decided_at":   "",
                "decided_by":   "",
            })
            _save_paid_leave(data)
            refresh_request_tree()
            rq_note_var.set("")
            self.set_status(f"「{name}」{rq_date_var.get().strip()} の有給申請を受け付けました。")

        def _get_selected_request():
            sel = request_tree.selection()
            if not sel:
                messagebox.showwarning("選択", "対象の申請をリストから選択してください。", parent=dlg)
                return None
            vals = request_tree.item(sel[0], "values")
            data = _load_paid_leave()
            for r in data["requests"]:
                if str(r.get("id", "")) == str(vals[0]):
                    return data, r
            return None

        def approve_request():
            found = _get_selected_request()
            if not found:
                return
            data, req = found
            if req.get("status") != "申請中":
                messagebox.showwarning("承認", "この申請はすでに処理済みです。", parent=dlg)
                return
            rem, _, _ = _calc_paid_leave_balance(req["name"], data)
            if float(req["days"]) > rem:
                if not messagebox.askyesno(
                        "残日数不足",
                        f"「{req['name']}」の残日数({rem}日)を超えています。\n承認しますか？", parent=dlg):
                    return
            data["usages"].append({
                "id":   _next_pl_id(data["usages"]),
                "name": req["name"],
                "date": req["leave_date"],
                "days": req["days"],
                "note": f"申請承認（{req.get('note', '')}）" if req.get("note") else "申請承認",
            })
            req["status"]     = "承認"
            req["decided_at"] = datetime.date.today().isoformat()
            req["decided_by"] = self.logged_in_user or ""
            _save_paid_leave(data)
            refresh_request_tree()
            refresh_grant_tree()
            refresh_usage_tree()
            refresh_balance()
            update_remain_lbl()
            self.set_status(f"「{req['name']}」の有給申請を承認しました。")

        def reject_request():
            found = _get_selected_request()
            if not found:
                return
            data, req = found
            if req.get("status") != "申請中":
                messagebox.showwarning("却下", "この申請はすでに処理済みです。", parent=dlg)
                return
            if not messagebox.askyesno("確認", f"「{req['name']}」の申請を却下しますか？", parent=dlg):
                return
            req["status"]     = "却下"
            req["decided_at"] = datetime.date.today().isoformat()
            req["decided_by"] = self.logged_in_user or ""
            _save_paid_leave(data)
            refresh_request_tree()
            self.set_status(f"「{req['name']}」の有給申請を却下しました。")

        def delete_request():
            found = _get_selected_request()
            if not found:
                return
            data, req = found
            if req.get("status") == "申請中":
                messagebox.showwarning("削除", "申請中の項目は先に承認または却下してください。", parent=dlg)
                return
            if not messagebox.askyesno("確認", f"ID:{req['id']} の申請履歴を削除しますか？", parent=dlg):
                return
            data["requests"] = [r for r in data["requests"] if str(r.get("id", "")) != str(req["id"])]
            _save_paid_leave(data)
            refresh_request_tree()

        btn_rq = tk.Frame(tab_request, bg=C_BG)
        btn_rq.pack(pady=(2, 4))
        _color_btn(btn_rq, "申請する", submit_request, C_GRANT, width=10).pack(side=tk.LEFT, padx=4)
        if is_admin:
            _color_btn(btn_rq, "承認", approve_request, C_BTN_CHECK, width=8).pack(side=tk.LEFT, padx=4)
            _color_btn(btn_rq, "却下", reject_request,  C_BTN_DEL,   width=8).pack(side=tk.LEFT, padx=4)
            _color_btn(btn_rq, "削除", delete_request,  C_BTN_REFR,  width=8).pack(side=tk.LEFT, padx=4)
        request_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        sb_request.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 8))
        refresh_request_tree()

        # ── タブ2: 付与登録 ──────────────────────
        gf = tk.Frame(tab_grant, bg=C_BG)
        gf.pack(fill=tk.X, padx=12, pady=10)

        def glbl(text, row, col):
            tk.Label(gf, text=text, bg=C_BG, font=FONT_MAIN).grid(
                row=row, column=col, sticky=tk.W, padx=(0, 4), pady=4)

        glbl("スタッフ名:", 0, 0)
        g_name_var = tk.StringVar()
        ttk.Combobox(gf, textvariable=g_name_var,
                     values=[s.get("name") for s in self.staff_list if isinstance(s, dict)],
                     width=16, state="readonly").grid(row=0, column=1, sticky=tk.W)
        glbl("付与日:", 0, 2)
        g_date_var = tk.StringVar(value=datetime.date.today().isoformat())
        ttk.Entry(gf, textvariable=g_date_var, width=14).grid(row=0, column=3, sticky=tk.W)
        glbl("付与日数:", 1, 0)
        g_days_var = tk.StringVar(value="10")
        ttk.Entry(gf, textvariable=g_days_var, width=8).grid(row=1, column=1, sticky=tk.W)
        glbl("備考:", 2, 0)
        g_note_var = tk.StringVar()
        ttk.Entry(gf, textvariable=g_note_var, width=36).grid(
            row=2, column=1, columnspan=3, sticky=tk.W + tk.E)

        grant_cols = ("id", "name", "grant_date", "days", "expiry", "note")
        grant_tree = ttk.Treeview(tab_grant, columns=grant_cols, show="headings", height=11)
        for cid, ht, w, anch in [
            ("id",         "ID",       50,  tk.CENTER),
            ("name",       "スタッフ名", 130, tk.W),
            ("grant_date", "付与日",   100, tk.CENTER),
            ("days",       "付与日数",  80, tk.CENTER),
            ("expiry",     "有効期限", 100, tk.CENTER),
            ("note",       "備考",     210, tk.W),
        ]:
            grant_tree.heading(cid, text=ht)
            grant_tree.column(cid, width=w, anchor=anch)
        grant_tree.tag_configure("odd",  background=C_ROW_ODD)
        grant_tree.tag_configure("even", background=C_ROW_EVEN)
        sb_grant = ttk.Scrollbar(tab_grant, orient=tk.VERTICAL, command=grant_tree.yview)
        grant_tree.configure(yscroll=sb_grant.set)

        def refresh_grant_tree():
            data = _load_paid_leave()
            for it in grant_tree.get_children():
                grant_tree.delete(it)
            for i, g in enumerate(sorted(data["grants"],
                                         key=lambda x: x.get("grant_date", ""), reverse=True)):
                tag = "even" if i % 2 == 0 else "odd"
                expiry = g.get("expiry_date", "")
                expiry_disp = "無期限" if expiry in ("", "9999-12-31") else expiry
                grant_tree.insert("", tk.END, tags=(tag,), values=(
                    g.get("id", ""), g.get("name", ""), g.get("grant_date", ""),
                    g.get("days", ""), expiry_disp, g.get("note", "")))

        def add_grant():
            name = g_name_var.get().strip()
            if not name:
                messagebox.showwarning("入力エラー", "スタッフ名を選択してください。", parent=dlg)
                return
            try:
                datetime.date.fromisoformat(g_date_var.get().strip())
                days = float(g_days_var.get().strip())
                if days <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "付与日と付与日数（正の数）を正しく入力してください。", parent=dlg)
                return
            data = _load_paid_leave()
            data["grants"].append({
                "id":          _next_pl_id(data["grants"]),
                "name":        name,
                "grant_date":  g_date_var.get().strip(),
                "days":        days,
                "expiry_date": "9999-12-31",  # 有効期限は廃止し、常に無期限として扱う
                "note":        g_note_var.get().strip(),
            })
            _save_paid_leave(data)
            refresh_grant_tree()
            refresh_balance()
            self.set_status(f"「{name}」に {days}日 の有給を付与しました。")

        def delete_grant():
            sel = grant_tree.selection()
            if not sel:
                messagebox.showwarning("削除", "削除する付与レコードを選択してください。", parent=dlg)
                return
            vals = grant_tree.item(sel[0], "values")
            if not messagebox.askyesno("確認", f"ID:{vals[0]} の付与レコードを削除しますか？", parent=dlg):
                return
            data = _load_paid_leave()
            data["grants"] = [g for g in data["grants"] if str(g.get("id", "")) != str(vals[0])]
            _save_paid_leave(data)
            refresh_grant_tree()
            refresh_balance()

        btn_gf = tk.Frame(tab_grant, bg=C_BG)
        btn_gf.pack(pady=(2, 4))
        _color_btn(btn_gf, "付与登録", add_grant,    C_GRANT,    width=10).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_gf, "削除",     delete_grant, C_BTN_DEL,  width=8).pack(side=tk.LEFT, padx=4)
        grant_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        sb_grant.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 8))
        refresh_grant_tree()

        # ── タブ3: 取得登録 ──────────────────────
        uf = tk.Frame(tab_usage, bg=C_BG)
        uf.pack(fill=tk.X, padx=12, pady=10)

        def ulbl(text, row, col):
            tk.Label(uf, text=text, bg=C_BG, font=FONT_MAIN).grid(
                row=row, column=col, sticky=tk.W, padx=(0, 4), pady=4)

        ulbl("スタッフ名:", 0, 0)
        u_name_var = tk.StringVar()
        ttk.Combobox(uf, textvariable=u_name_var,
                     values=[s.get("name") for s in self.staff_list if isinstance(s, dict)],
                     width=16, state="readonly").grid(row=0, column=1, sticky=tk.W)
        ulbl("取得日:", 0, 2)
        u_date_var = tk.StringVar(value=datetime.date.today().isoformat())
        ttk.Entry(uf, textvariable=u_date_var, width=14).grid(row=0, column=3, sticky=tk.W)
        ulbl("取得日数:", 1, 0)
        u_days_var = tk.StringVar(value="1")
        ttk.Entry(uf, textvariable=u_days_var, width=8).grid(row=1, column=1, sticky=tk.W)
        tk.Label(uf, text="(0.5=半日)", bg=C_BG, font=("Meiryo UI", 9)).grid(
            row=1, column=2, sticky=tk.W)
        ulbl("備考:", 2, 0)
        u_note_var = tk.StringVar()
        ttk.Entry(uf, textvariable=u_note_var, width=36).grid(
            row=2, column=1, columnspan=3, sticky=tk.W + tk.E)
        u_remain_var = tk.StringVar(value="残日数: - 日")
        tk.Label(uf, textvariable=u_remain_var, bg=C_BG, font=FONT_BOLD,
                 fg=C_GRANT).grid(row=0, column=4, padx=(16, 0), sticky=tk.W)

        def update_remain_lbl(*_):
            name = u_name_var.get().strip()
            if name:
                rem, _, _ = _calc_paid_leave_balance(name, _load_paid_leave())
                u_remain_var.set(f"残日数: {rem} 日")
            else:
                u_remain_var.set("残日数: - 日")
        u_name_var.trace_add("write", update_remain_lbl)

        usage_cols = ("id", "name", "date", "days", "note")
        usage_tree = ttk.Treeview(tab_usage, columns=usage_cols, show="headings", height=11)
        for cid, ht, w, anch in [
            ("id",   "ID",        50,  tk.CENTER),
            ("name", "スタッフ名", 130, tk.W),
            ("date", "取得日",    100, tk.CENTER),
            ("days", "取得日数",   80, tk.CENTER),
            ("note", "備考",      310, tk.W),
        ]:
            usage_tree.heading(cid, text=ht)
            usage_tree.column(cid, width=w, anchor=anch)
        usage_tree.tag_configure("odd",  background=C_ROW_ODD)
        usage_tree.tag_configure("even", background=C_ROW_EVEN)
        sb_usage = ttk.Scrollbar(tab_usage, orient=tk.VERTICAL, command=usage_tree.yview)
        usage_tree.configure(yscroll=sb_usage.set)

        def refresh_usage_tree():
            data = _load_paid_leave()
            for it in usage_tree.get_children():
                usage_tree.delete(it)
            for i, u in enumerate(sorted(data["usages"],
                                         key=lambda x: x.get("date", ""), reverse=True)):
                tag = "even" if i % 2 == 0 else "odd"
                usage_tree.insert("", tk.END, tags=(tag,), values=(
                    u.get("id", ""), u.get("name", ""), u.get("date", ""),
                    u.get("days", ""), u.get("note", "")))

        def add_usage():
            name = u_name_var.get().strip()
            if not name:
                messagebox.showwarning("入力エラー", "スタッフ名を選択してください。", parent=dlg)
                return
            try:
                datetime.date.fromisoformat(u_date_var.get().strip())
                days = float(u_days_var.get().strip())
                if days <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "取得日と取得日数（正の数）を正しく入力してください。", parent=dlg)
                return
            data = _load_paid_leave()
            rem, _, _ = _calc_paid_leave_balance(name, data)
            if days > rem:
                if not messagebox.askyesno(
                        "残日数不足",
                        f"残日数({rem}日)を超えています。\n続けますか？", parent=dlg):
                    return
            data["usages"].append({
                "id":   _next_pl_id(data["usages"]),
                "name": name,
                "date": u_date_var.get().strip(),
                "days": days,
                "note": u_note_var.get().strip(),
            })
            _save_paid_leave(data)
            refresh_usage_tree()
            refresh_balance()
            update_remain_lbl()
            self.set_status(f"「{name}」{u_date_var.get().strip()} 有給{days}日を登録しました。")

        def delete_usage():
            sel = usage_tree.selection()
            if not sel:
                messagebox.showwarning("削除", "削除する取得レコードを選択してください。", parent=dlg)
                return
            vals = usage_tree.item(sel[0], "values")
            if not messagebox.askyesno("確認", f"ID:{vals[0]} の取得レコードを削除しますか？", parent=dlg):
                return
            data = _load_paid_leave()
            data["usages"] = [u for u in data["usages"] if str(u.get("id", "")) != str(vals[0])]
            _save_paid_leave(data)
            refresh_usage_tree()
            refresh_balance()
            update_remain_lbl()

        btn_uf = tk.Frame(tab_usage, bg=C_BG)
        btn_uf.pack(pady=(2, 4))
        _color_btn(btn_uf, "取得登録", add_usage,    C_GRANT,   width=10).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_uf, "削除",     delete_usage, C_BTN_DEL, width=8).pack(side=tk.LEFT, padx=4)
        usage_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        sb_usage.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 8))
        refresh_usage_tree()

        # ── タブ4: 履歴一覧 ──────────────────────
        hist_ctrl = tk.Frame(tab_history, bg=C_BG)
        hist_ctrl.pack(fill=tk.X, padx=8, pady=6)
        tk.Label(hist_ctrl, text="スタッフ:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        hist_name_var = tk.StringVar(value="(全員)")
        ttk.Combobox(hist_ctrl, textvariable=hist_name_var,
                     values=["(全員)"] + [s.get("name") for s in self.staff_list if isinstance(s, dict)],
                     width=14, state="readonly").pack(side=tk.LEFT, padx=(4, 10))
        tk.Label(hist_ctrl, text="年:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        hist_year_var = tk.StringVar(value=str(datetime.date.today().year))
        ttk.Entry(hist_ctrl, textvariable=hist_year_var, width=7).pack(side=tk.LEFT, padx=(4, 8))

        hist_cols = ("kind", "name", "date", "days", "note")
        hist_tree = ttk.Treeview(tab_history, columns=hist_cols, show="headings", height=14)
        for cid, ht, w, anch in [
            ("kind", "種別",       70,  tk.CENTER),
            ("name", "スタッフ名", 130, tk.W),
            ("date", "日付",      100, tk.CENTER),
            ("days", "日数",       80, tk.CENTER),
            ("note", "備考",      290, tk.W),
        ]:
            hist_tree.heading(cid, text=ht)
            hist_tree.column(cid, width=w, anchor=anch)
        hist_tree.tag_configure("grant", background="#D5F5E3")
        hist_tree.tag_configure("usage", background="#FDEBD0")
        sb_hist = ttk.Scrollbar(tab_history, orient=tk.VERTICAL, command=hist_tree.yview)
        hist_tree.configure(yscroll=sb_hist.set)

        def refresh_history():
            nf   = hist_name_var.get().strip()
            yf   = hist_year_var.get().strip()
            data = _load_paid_leave()
            for it in hist_tree.get_children():
                hist_tree.delete(it)
            entries = []
            for g in data["grants"]:
                if (nf == "(全員)" or g.get("name") == nf) and \
                        (not yf or g.get("grant_date", "").startswith(yf)):
                    entries.append(("付与", g["name"], g.get("grant_date", ""),
                                    g.get("days", ""), g.get("note", "")))
            for u in data["usages"]:
                if (nf == "(全員)" or u.get("name") == nf) and \
                        (not yf or u.get("date", "").startswith(yf)):
                    entries.append(("取得", u["name"], u.get("date", ""),
                                    u.get("days", ""), u.get("note", "")))
            entries.sort(key=lambda x: x[2], reverse=True)
            for rec in entries:
                hist_tree.insert("", tk.END, tags=("grant" if rec[0] == "付与" else "usage",),
                                 values=rec)

        _color_btn(hist_ctrl, "表示", refresh_history, C_BTN_REFR, width=8).pack(side=tk.LEFT)
        hist_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        sb_hist.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 8))
        refresh_history()

    # ── 給与計算 ──────────────────────────────────
    def open_payroll_manager(self):
        dlg = tk.Toplevel(self)
        dlg.title("給与計算")
        dlg.geometry("1120x580")
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        tk.Label(
            dlg,
            text="⚠ 概算計算です。社会保険料は簡易計算（標準報酬月額等級表は使用せず、当月の総支給額に料率を"
                 "乗じています）、所得税(乙欄)は簡易近似、住民税は手入力です。"
                 "実際の給与支払前に社会保険労務士・税理士等の確認を必ず受けてください。",
            bg="#FFF3CD", fg="#7A5C00", font=("Meiryo UI", 9), wraplength=970,
            justify=tk.LEFT, anchor=tk.W,
        ).pack(fill=tk.X, padx=10, pady=(10, 4))

        ctrl = tk.Frame(dlg, bg=C_BG)
        ctrl.pack(fill=tk.X, padx=10, pady=4)
        tk.Label(ctrl, text="年:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        year_var = tk.StringVar(value=str(datetime.date.today().year))
        ttk.Entry(ctrl, textvariable=year_var, width=6).pack(side=tk.LEFT, padx=(2, 10))
        tk.Label(ctrl, text="月:", bg=C_BG, font=FONT_MAIN).pack(side=tk.LEFT)
        month_var = tk.StringVar(value=str(datetime.date.today().month))
        ttk.Combobox(ctrl, textvariable=month_var, values=[str(m) for m in range(1, 13)],
                     width=4, state="readonly").pack(side=tk.LEFT, padx=(2, 10))

        results_holder = {"data": []}

        def yen(v):
            return f"{int(round(v)):,}"

        cols = ("name", "total_hours", "gross", "social", "income_tax", "resident_tax",
                "deductions", "adjustment", "net")
        tree = ttk.Treeview(dlg, columns=cols, show="headings", height=16)
        headers = {
            "name":         ("スタッフ名",     110, tk.W),
            "total_hours":  ("総労働時間(h)",   90, tk.CENTER),
            "gross":        ("総支給額",       100, tk.E),
            "social":       ("社会保険料計",   100, tk.E),
            "income_tax":   ("所得税",          90, tk.E),
            "resident_tax": ("住民税",          90, tk.E),
            "deductions":   ("控除合計",       100, tk.E),
            "adjustment":   ("その他調整額",   100, tk.E),
            "net":          ("差引支給額",     110, tk.E),
        }
        for cid, (ht, w, anch) in headers.items():
            tree.heading(cid, text=ht)
            tree.column(cid, width=w, anchor=anch)
        tree.tag_configure("odd",  background=C_ROW_ODD)
        tree.tag_configure("even", background=C_ROW_EVEN)
        tree.tag_configure("shortfall", background="#FADBD8")

        def run_calc():
            try:
                year = int(year_var.get().strip())
                month = int(month_var.get().strip())
                if not (1 <= month <= 12):
                    raise ValueError
            except ValueError:
                messagebox.showwarning("入力エラー", "年・月を正しく入力してください。", parent=dlg)
                return
            am = labor.AttendanceManager()
            settings = payroll.load_payroll_settings()
            results = payroll.compute_payroll_for_all(self.staff_list, am, year, month, settings)
            results_holder["data"] = results
            for it in tree.get_children():
                tree.delete(it)
            shortfall_names = []
            for i, res in enumerate(results):
                shortfall = res.get("net_pay_shortfall", 0)
                if shortfall > 0:
                    tag = "shortfall"
                    shortfall_names.append(f"{res['name']}（{yen(shortfall)}円不足）")
                    net_text = f"0 ⚠-{yen(shortfall)}"
                else:
                    tag = "even" if i % 2 == 0 else "odd"
                    net_text = yen(res["net_pay"])
                tree.insert("", tk.END, tags=(tag,), values=(
                    res["name"],
                    res["breakdown"]["total_hours"],
                    yen(res["gross_total"]),
                    yen(res["social_insurance_total"]),
                    yen(res["income_tax"]),
                    yen(res["resident_tax"]),
                    yen(res["total_deductions"]),
                    yen(res["adjustment_amount"]),
                    net_text,
                ))
            base_msg = f"{year}年{month}月分の給与計算を実行しました（{len(results)}名、時給未設定=0円のスタッフを含む）。"
            if shortfall_names:
                self.set_status(
                    base_msg + f" ⚠ 調整額により差引支給額が0円未満になったため0円に補正: {'、'.join(shortfall_names)}",
                    error=True,
                )
            else:
                self.set_status(base_msg)

        def on_tree_double_click(event=None):
            sel = tree.selection()
            if not sel:
                return
            idx = tree.index(sel[0])
            if idx >= len(results_holder["data"]):
                return
            res = results_holder["data"][idx]
            bd = res["breakdown"]
            gross = res["gross"]
            wage_line = (f"給与形態: {res['pay_type']}　"
                         + (f"換算時給: {yen(gross.get('premium_wage', 0))}円"
                            if res.get("pay_type") == "月給" else f"時給: {yen(res['hourly_wage'])}円"))
            adjustment_line = f"その他調整額: {yen(res['adjustment_amount'])}円"
            if res.get("adjustment_note"):
                adjustment_line += f"（{res['adjustment_note']}）"
            net_pay_line = f"差引支給額: {yen(res['net_pay'])}円"
            if res.get("net_pay_shortfall", 0) > 0:
                net_pay_line += (
                    f"\n⚠ 調整額が大きいため本来は {yen(res['net_pay_raw'])}円 "
                    f"（{yen(res['net_pay_shortfall'])}円不足）となり、0円に補正して表示しています。"
                )
            detail = (
                f"【{res['name']}　{res['year']}年{res['month']}月分】\n\n"
                f"{wage_line}\n"
                f"─ 労働時間内訳 ─\n"
                f"通常: {bd['regular_hours']}h　残業25%: {bd['overtime25_hours']}h　"
                f"残業50%: {bd['overtime50_hours']}h\n"
                f"法定休日: {bd['holiday_hours']}h　深夜加算対象: {bd['night_hours']}h\n"
                f"総労働時間: {bd['total_hours']}h\n\n"
                f"【支給】\n"
                f"基本給: {yen(res['base_pay_amount'])}円\n"
                f"時給: {yen(res['hourly_pay_amount'])}円\n"
                f"残業25%: {yen(gross['pay_overtime25'])}円\n"
                f"残業50%: {yen(gross['pay_overtime50'])}円\n"
                f"休日出勤35%: {yen(gross['pay_holiday'])}円\n"
                f"深夜割増加算: {yen(gross['pay_night_addon'])}円\n"
                f"役職手当: {yen(res['position_allowance'])}円\n"
                f"皆勤手当: {yen(res['attendance_allowance'])}円\n"
                f"住宅手当: {yen(res['housing_allowance'])}円\n"
                f"交通費: {yen(res['commute_allowance'])}円\n"
                f"総支給額: {yen(res['gross_total'])}円\n\n"
                f"【控除】\n"
                f"健康保険: {yen(res['health_insurance'])}円\n"
                f"厚生年金: {yen(res['pension'])}円\n"
                f"介護保険: {yen(res['care_insurance'])}円\n"
                f"子育て支援金: {yen(res['child_support_levy'])}円\n"
                f"雇用保険: {yen(res['employment_insurance'])}円\n"
                f"社会保険料(計): {yen(res['social_insurance_total'])}円\n"
                f"所得税({res['tax_table']}欄): {yen(res['income_tax'])}円\n"
                f"住民税: {yen(res['resident_tax'])}円\n"
                f"控除合計: {yen(res['total_deductions'])}円\n\n"
                f"{adjustment_line}\n\n"
                f"{net_pay_line}"
            )
            messagebox.showinfo("給与明細", detail, parent=dlg)

        tree.bind("<Double-1>", on_tree_double_click)

        def open_settings():
            self._open_payroll_settings_dialog(dlg)

        def edit_adjustment():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("調整額編集", "対象のスタッフをリストから選択してください。", parent=dlg)
                return
            idx = tree.index(sel[0])
            if idx >= len(results_holder["data"]):
                return
            res = results_holder["data"][idx]
            self._open_adjustment_dialog(dlg, res["name"], res["year"], res["month"],
                                          on_saved=run_calc)

        def export_payroll_excel():
            if not results_holder["data"]:
                messagebox.showwarning("Excel出力", "先に「計算」を実行してください。", parent=dlg)
                return
            self._export_payroll_excel(
                results_holder["data"], int(year_var.get()), int(month_var.get()))

        def export_payslips():
            if not results_holder["data"]:
                messagebox.showwarning("給与明細出力", "先に「計算」を実行してください。", parent=dlg)
                return
            self._export_individual_payslips(
                results_holder["data"], int(year_var.get()), int(month_var.get()))

        def export_payslips_pdf():
            if not results_holder["data"]:
                messagebox.showwarning("給与明細PDF出力", "先に「計算」を実行してください。", parent=dlg)
                return
            self._export_individual_payslips_pdf(
                results_holder["data"], int(year_var.get()), int(month_var.get()))

        btn_ctrl = tk.Frame(dlg, bg=C_BG)
        btn_ctrl.pack(fill=tk.X, padx=10, pady=(0, 4))
        _color_btn(btn_ctrl, "計算",       run_calc,            C_BTN_CHECK, width=10).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_ctrl, "✏ 調整額編集", edit_adjustment,     "#8E44AD",   width=12).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_ctrl, "⚙ 料率設定", open_settings,        C_BTN_REFR,  width=12).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_ctrl, "📥 Excel出力", export_payroll_excel, C_BTN_XLS,   width=12).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_ctrl, "📄 給与明細出力", export_payslips,   "#2E86C1",   width=14).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_ctrl, "🧾 給与明細PDF出力", export_payslips_pdf, "#C0392B", width=16).pack(side=tk.LEFT, padx=4)
        tk.Label(btn_ctrl, text="（行をダブルクリックで明細表示、選択して調整額編集も可能）", bg=C_BG,
                 font=("Meiryo UI", 8)).pack(side=tk.LEFT, padx=(10, 0))

        tree_frame = tk.Frame(dlg, bg=C_BG)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        sb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=sb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        run_calc()

    def _open_adjustment_dialog(self, parent, name, year, month, on_saved=None):
        current = payroll.get_adjustment(name, year, month)
        dlg = tk.Toplevel(parent)
        dlg.title(f"その他調整額 - {name}（{year}年{month}月）")
        dlg.geometry("380x260")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        tk.Label(
            dlg,
            text="この月だけの個別調整額です。プラス（増額）・マイナス（減額）どちらも\n"
                 "入力できます。税額・社会保険料の計算には影響せず、差引支給額に\n"
                 "直接加減算されます。",
            bg=C_BG, font=("Meiryo UI", 9), justify=tk.LEFT,
        ).pack(padx=14, pady=(14, 8), anchor=tk.W)

        f = tk.Frame(dlg, bg=C_BG)
        f.pack(fill=tk.X, padx=14)
        tk.Label(f, text="調整額(円):", bg=C_BG, font=FONT_MAIN).grid(row=0, column=0, sticky=tk.W, pady=4)
        amount_var = tk.StringVar(value=str(int(current["amount"])) if current["amount"] else "0")
        ttk.Entry(f, textvariable=amount_var, width=14).grid(row=0, column=1, sticky=tk.W)
        tk.Label(f, text="※マイナスの場合は先頭に「-」を付けてください（例: -3000）", bg=C_BG,
                 font=("Meiryo UI", 8), justify=tk.LEFT).grid(row=1, column=0, columnspan=2, sticky=tk.W)

        tk.Label(f, text="理由・メモ:", bg=C_BG, font=FONT_MAIN).grid(row=2, column=0, sticky=tk.W, pady=(10, 4))
        note_var = tk.StringVar(value=current["note"])
        ttk.Entry(f, textvariable=note_var, width=32).grid(row=2, column=1, sticky=tk.W, pady=(10, 4))

        def save():
            try:
                amount = float(amount_var.get().strip() or 0)
            except ValueError:
                messagebox.showwarning("入力エラー", "調整額は数値で入力してください。", parent=dlg)
                return
            payroll.set_adjustment(name, year, month, amount, note_var.get().strip())
            self.set_status(f"「{name}」{year}年{month}月分のその他調整額を保存しました。")
            if on_saved:
                on_saved()
            dlg.destroy()

        def clear():
            payroll.set_adjustment(name, year, month, 0, "")
            self.set_status(f"「{name}」{year}年{month}月分のその他調整額をクリアしました。")
            if on_saved:
                on_saved()
            dlg.destroy()

        btn_f = tk.Frame(dlg, bg=C_BG)
        btn_f.pack(pady=(16, 10))
        _color_btn(btn_f, "保存",     save,        C_BTN_CHECK, width=8).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_f, "クリア",   clear,       C_BTN_DEL,   width=8).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_f, "閉じる",   dlg.destroy, C_BTN_REFR,  width=8).pack(side=tk.LEFT, padx=4)

    def _open_payroll_settings_dialog(self, parent):
        settings = payroll.load_payroll_settings()
        dlg = tk.Toplevel(parent)
        dlg.title("給与計算 - 料率設定")
        dlg.geometry("480x560")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()

        tk.Label(
            dlg,
            text="保険料率・税額計算に関する設定です。年度・都道府県・業種により変わるため、\n"
                 "最新の公的情報（協会けんぽ・日本年金機構・厚生労働省・国税庁）に基づき\n"
                 "必要に応じて更新してください。",
            bg=C_BG, font=("Meiryo UI", 9), justify=tk.LEFT,
        ).pack(padx=14, pady=(12, 6), anchor=tk.W)

        fields = [
            ("overtime_rate_25",                   "時間外労働 割増率(%)"),
            ("overtime_rate_50",                   "月60時間超 時間外割増率(%)"),
            ("overtime_monthly_threshold_hours",   "月60時間超の閾値(時間)"),
            ("holiday_rate_35",                    "法定休日労働 割増率(%)"),
            ("night_rate_25",                      "深夜労働(22-5時) 加算率(%)"),
            ("health_insurance_rate_employee",     "健康保険料率(従業員負担,%)"),
            ("care_insurance_rate_employee",       "介護保険料率(40-64歳,従業員負担,%)"),
            ("pension_rate_employee",              "厚生年金保険料率(従業員負担,%)"),
            ("employment_insurance_rate_employee", "雇用保険料率(従業員負担,%)"),
            ("child_support_levy_rate_employee",   "子ども・子育て支援金等(従業員負担,%)"),
            ("income_tax_otsu_rate_percent",       "所得税(乙欄)簡易近似率(%)"),
            ("standard_monthly_hours",             "月給制の所定労働時間(月間,時間)"),
        ]

        f = tk.Frame(dlg, bg=C_BG)
        f.pack(fill=tk.BOTH, expand=True, padx=14)
        vars_map = {}
        for i, (key, label) in enumerate(fields):
            tk.Label(f, text=label, bg=C_BG, font=("Meiryo UI", 9), wraplength=280,
                     justify=tk.LEFT).grid(row=i, column=0, sticky=tk.W, pady=5)
            v = tk.StringVar(value=str(settings.get(key, "")))
            ttk.Entry(f, textvariable=v, width=10).grid(row=i, column=1, sticky=tk.E, pady=5)
            vars_map[key] = v

        def save():
            new_settings = dict(settings)
            for key, v in vars_map.items():
                try:
                    new_settings[key] = float(v.get().strip())
                except ValueError:
                    messagebox.showwarning("入力エラー", f"数値を入力してください（{key}）。", parent=dlg)
                    return
            payroll.save_payroll_settings(new_settings)
            self.set_status("給与計算の料率設定を保存しました。")
            dlg.destroy()

        btn_f = tk.Frame(dlg, bg=C_BG)
        btn_f.pack(pady=12)
        _color_btn(btn_f, "保存",   save,        C_BTN_CHECK, width=10).pack(side=tk.LEFT, padx=4)
        _color_btn(btn_f, "閉じる", dlg.destroy, C_BTN_REFR,  width=10).pack(side=tk.LEFT, padx=4)

    def _export_payroll_excel(self, results, year, month):
        default_name = f"給与計算_{year}-{month:02d}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Excel ファイルの保存先",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx")],
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "給与計算"

        hdr_fill  = PatternFill("solid", fgColor="8E44AD")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Meiryo UI", size=10)
        body_font = Font(name="Meiryo UI", size=10)
        even_fill = PatternFill("solid", fgColor="EAF0FB")
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center")
        thin      = Side(style="thin", color="CCCCCC")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        title_cell = ws.cell(row=1, column=1, value=f"給与計算 {year}年{month}月分（概算・要確認）")
        title_cell.font = Font(bold=True, size=12, name="Meiryo UI")

        headers = [
            "スタッフ名", "総労働時間(h)", "通常(h)", "残業25%(h)", "残業50%(h)", "法定休日(h)", "深夜加算(h)",
            # 【支給】
            "基本給", "時給", "残業25%", "残業50%", "休日出勤", "深夜手当",
            "役職手当", "皆勤手当", "住宅手当", "交通費", "総支給額",
            # 【控除】
            "健康保険", "厚生年金", "介護保険", "子育て支援金", "雇用保険", "社会保険料",
            "所得税", "住民税", "控除合計",
            "その他調整額", "調整理由", "差引支給額",
        ]
        col_widths = [
            14, 12, 10, 10, 10, 10, 10,
            11, 11, 10, 10, 10, 10, 10, 10, 10, 10, 12,
            10, 10, 10, 12, 10, 12, 10, 10, 12,
            12, 20, 12,
        ]
        ncols = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        for ri, res in enumerate(results, 3):
            bd = res["breakdown"]
            gross = res["gross"]
            fill = even_fill if ri % 2 == 0 else PatternFill()
            vals = [
                res["name"], bd["total_hours"], bd["regular_hours"], bd["overtime25_hours"],
                bd["overtime50_hours"], bd["holiday_hours"], bd["night_hours"],
                res["base_pay_amount"], res["hourly_pay_amount"], gross["pay_overtime25"],
                gross["pay_overtime50"], gross["pay_holiday"], gross["pay_night_addon"],
                res["position_allowance"], res["attendance_allowance"], res["housing_allowance"],
                res["commute_allowance"], res["gross_total"],
                res["health_insurance"], res["pension"], res["care_insurance"],
                res["child_support_levy"], res["employment_insurance"], res["social_insurance_total"],
                res["income_tax"], res["resident_tax"], res["total_deductions"],
                res["adjustment_amount"], res["adjustment_note"], res["net_pay"],
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = body_font; c.fill = fill; c.border = border
                c.alignment = left if ci == 1 else center
        ws.freeze_panes = "A3"

        note_row = len(results) + 4
        note = ws.cell(
            row=note_row, column=1,
            value="※概算計算です。社会保険料は簡易計算（標準報酬月額等級表は未使用）、"
                  "所得税(乙欄)は簡易近似です。交通費は非課税として所得税の課税対象額から除外しています。"
                  "実際の給与支払前に社会保険労務士・税理士等の確認を必ず受けてください。",
        )
        note.font = Font(italic=True, color="B26A00", name="Meiryo UI", size=9)
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)

        try:
            wb.save(path)
            messagebox.showinfo("完了", f"Excelファイルを保存しました。\n{path}")
        except Exception as exc:
            messagebox.showerror("エラー", f"保存に失敗しました: {exc}")

    def _export_individual_payslips(self, results, year, month):
        """スタッフ1人につき1シートの、個別の給与明細書を出力する。"""
        default_name = f"給与明細_{year}-{month:02d}.xlsx"
        path = filedialog.asksaveasfilename(
            title="給与明細 Excel ファイルの保存先",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx")],
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        title_font  = Font(bold=True, size=16, name="Meiryo UI")
        sub_font    = Font(size=10, name="Meiryo UI", color="555555")
        label_font  = Font(bold=True, size=10, name="Meiryo UI")
        value_font  = Font(size=10, name="Meiryo UI")
        section_font = Font(bold=True, size=10, color="FFFFFF", name="Meiryo UI")
        total_font  = Font(bold=True, size=13, name="Meiryo UI")
        note_font   = Font(italic=True, size=8, color="888888", name="Meiryo UI")
        section_fill = PatternFill("solid", fgColor="4A7FD4")
        total_fill   = PatternFill("solid", fgColor="D5E8D4")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")
        right  = Alignment(horizontal="right",  vertical="center")

        def cell(ws, row, col, value=None, font=None, fill=None, align=None, brd=None):
            c = ws.cell(row=row, column=col, value=value)
            if font:  c.font = font
            if fill:  c.fill = fill
            if align: c.alignment = align
            if brd:   c.border = brd
            return c

        def yen(v):
            return f"{int(round(v)):,}円"

        used_titles = set()

        for res in results:
            raw_title = res["name"] or "無名"
            for ch in '[]:*?/\\':
                raw_title = raw_title.replace(ch, "")
            title = raw_title[:31] or "無名"
            n = 2
            while title in used_titles:
                suffix = f"({n})"
                title = (raw_title[:31 - len(suffix)] or "無名") + suffix
                n += 1
            used_titles.add(title)

            ws = wb.create_sheet(title=title)
            for col, w in zip("ABCDEF", (16, 14, 16, 16, 14, 4)):
                ws.column_dimensions[col].width = w

            cell(ws, 1, 1, "給与明細書", font=title_font, align=center)
            ws.merge_cells("A1:E1")
            cell(ws, 2, 1, f"{res['year']}年{res['month']}月分", font=sub_font, align=center)
            ws.merge_cells("A2:E2")

            cell(ws, 4, 1, "氏名:", font=label_font, align=left)
            cell(ws, 4, 2, f"{res['name']} 様", font=Font(bold=True, size=12, name="Meiryo UI"), align=left)
            ws.merge_cells("B4:C4")
            wage_label = "月給(基本給)" if res.get("pay_type") == "月給" else "時給"
            wage_value = yen(res['monthly_base_salary']) if res.get("pay_type") == "月給" else yen(res['hourly_wage'])
            cell(ws, 4, 4, f"{wage_label}:", font=label_font, align=left)
            cell(ws, 4, 5, wage_value, font=value_font, align=left)

            bd = res["breakdown"]
            cell(ws, 6, 1, "【勤怠】", font=section_font, fill=section_fill, align=left)
            ws.merge_cells("A6:E6")
            att_rows = [
                ("総労働時間(h)", bd["total_hours"], "法定休日(h)", bd["holiday_hours"]),
                ("残業25%(h)",   bd["overtime25_hours"], "残業50%(h)", bd["overtime50_hours"]),
                ("深夜加算対象(h)", bd["night_hours"], "", ""),
            ]
            r = 7
            for l1, v1, l2, v2 in att_rows:
                cell(ws, r, 1, l1, font=label_font, brd=border, align=left)
                cell(ws, r, 2, v1, font=value_font, brd=border, align=right)
                if l2:
                    cell(ws, r, 4, l2, font=label_font, brd=border, align=left)
                    cell(ws, r, 5, v2, font=value_font, brd=border, align=right)
                r += 1

            r += 1
            gross = res["gross"]
            cell(ws, r, 1, "【支給】", font=section_font, fill=section_fill, align=left)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell(ws, r, 4, "【控除】", font=section_font, fill=section_fill, align=left)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
            section_top = r + 1

            pay_items = [
                ("基本給",     res["base_pay_amount"]),
                ("時給",       res["hourly_pay_amount"]),
                ("残業25%",    gross["pay_overtime25"]),
                ("残業50%",    gross["pay_overtime50"]),
                ("休日出勤35%", gross["pay_holiday"]),
                ("深夜割増加算", gross["pay_night_addon"]),
                ("役職手当",   res["position_allowance"]),
                ("皆勤手当",   res["attendance_allowance"]),
                ("住宅手当",   res["housing_allowance"]),
                ("交通費",     res["commute_allowance"]),
            ]
            deduction_items = [
                ("健康保険",     res["health_insurance"]),
                ("厚生年金",     res["pension"]),
                ("介護保険",     res["care_insurance"]),
                ("子育て支援金", res["child_support_levy"]),
                ("雇用保険",     res["employment_insurance"]),
                ("社会保険料(計)", res["social_insurance_total"]),
                ("所得税",       res["income_tax"]),
                ("住民税",       res["resident_tax"]),
            ]

            for i, (label, val) in enumerate(pay_items):
                rr = section_top + i
                cell(ws, rr, 1, label, font=value_font, brd=border, align=left)
                cell(ws, rr, 2, yen(val), font=value_font, brd=border, align=right)
            for i, (label, val) in enumerate(deduction_items):
                rr = section_top + i
                is_bold = label == "社会保険料(計)"
                f = label_font if is_bold else value_font
                cell(ws, rr, 4, label, font=f, brd=border, align=left)
                cell(ws, rr, 5, yen(val), font=f, brd=border, align=right)

            items_bottom = section_top + max(len(pay_items), len(deduction_items))
            cell(ws, items_bottom, 1, "総支給額", font=label_font, fill=total_fill, brd=border, align=left)
            cell(ws, items_bottom, 2, yen(res["gross_total"]), font=label_font, fill=total_fill, brd=border, align=right)
            cell(ws, items_bottom, 4, "控除合計", font=label_font, fill=total_fill, brd=border, align=left)
            cell(ws, items_bottom, 5, yen(res["total_deductions"]), font=label_font, fill=total_fill, brd=border, align=right)

            r = items_bottom + 2
            cell(ws, r, 1, "その他調整額", font=label_font, align=left)
            cell(ws, r, 2, yen(res["adjustment_amount"]), font=value_font, align=right)
            if res.get("adjustment_note"):
                cell(ws, r, 3, f"（{res['adjustment_note']}）", font=note_font, align=left)
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

            r += 2
            cell(ws, r, 1, "差引支給額", font=total_font, fill=total_fill, align=left)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell(ws, r, 4, yen(res["net_pay"]), font=total_font, fill=total_fill, align=right)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

            r += 2
            cell(ws, r, 1,
                 "※概算計算です。社会保険料は簡易計算、所得税(乙欄)は簡易近似です。"
                 "支払前に社会保険労務士・税理士等の確認を受けてください。",
                 font=note_font, align=left)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

        try:
            wb.save(path)
            messagebox.showinfo("完了", f"給与明細（{len(results)}名分、スタッフごとにシート分割）を保存しました。\n{path}")
        except Exception as exc:
            messagebox.showerror("エラー", f"保存に失敗しました: {exc}")

    def _resolve_jp_pdf_font(self):
        """Windows標準フォントから日本語対応のTTF/TTCを探し、(通常, 太字)のパスを返す。
        見つからない場合は (None, None)。"""
        fonts_dir = r"C:\Windows\Fonts"
        candidates = [
            ("meiryo.ttc",   "meiryob.ttc"),
            ("msgothic.ttc", "msgothic.ttc"),
            ("YuGothR.ttc",  "YuGothB.ttc"),
            ("NotoSansJP-VF.ttf", "NotoSansJP-VF.ttf"),
        ]
        for regular, bold in candidates:
            r_path = os.path.join(fonts_dir, regular)
            b_path = os.path.join(fonts_dir, bold)
            if os.path.exists(r_path) and os.path.exists(b_path):
                return r_path, b_path
        return None, None

    def _export_individual_payslips_pdf(self, results, year, month):
        """スタッフ1人につき1ファイルの、個別の給与明細PDFを出力する。"""
        regular_path, bold_path = self._resolve_jp_pdf_font()
        if not regular_path:
            messagebox.showerror(
                "エラー",
                "日本語フォント（Meiryo / MS ゴシック等）が見つかりませんでした。\n"
                "Windows標準搭載のフォントが必要です。",
            )
            return

        folder = filedialog.askdirectory(title="給与明細PDFの保存先フォルダを選択")
        if not folder:
            return

        def yen(v):
            return f"{int(round(v)):,}円"

        def two_col_section(pdf, left_title, left_items, right_title, right_items):
            start_y = pdf.get_y()
            col_w = 88.0
            gap = 4.0
            left_x = pdf.l_margin
            right_x = left_x + col_w + gap

            pdf.set_font("JP", "B", 11)
            pdf.set_fill_color(74, 127, 212)
            pdf.set_text_color(255, 255, 255)
            pdf.set_xy(left_x, start_y)
            pdf.cell(col_w, 8, left_title, border=1, fill=True, align="C")
            pdf.set_xy(right_x, start_y)
            pdf.cell(col_w, 8, right_title, border=1, fill=True, align="C")
            pdf.set_text_color(0, 0, 0)

            y = start_y + 8
            pdf.set_font("JP", "", 10)
            max_rows = max(len(left_items), len(right_items))
            for i in range(max_rows):
                pdf.set_xy(left_x, y)
                if i < len(left_items):
                    label, val = left_items[i]
                    pdf.cell(col_w * 0.55, 7, label, border=1)
                    pdf.cell(col_w * 0.45, 7, val, border=1, align="R")
                else:
                    pdf.cell(col_w, 7, "", border=1)
                pdf.set_xy(right_x, y)
                if i < len(right_items):
                    label, val = right_items[i]
                    pdf.cell(col_w * 0.55, 7, label, border=1)
                    pdf.cell(col_w * 0.45, 7, val, border=1, align="R")
                else:
                    pdf.cell(col_w, 7, "", border=1)
                y += 7
            pdf.set_xy(left_x, y + 2)

        used_names = set()
        saved_paths = []

        for res in results:
            pdf = FPDF(unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.set_margins(15, 15, 15)
            pdf.add_font("JP", "", regular_path)
            pdf.add_font("JP", "B", bold_path)
            pdf.add_page()

            pdf.set_font("JP", "B", 20)
            pdf.cell(0, 12, "給与明細書", align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("JP", "", 11)
            pdf.cell(0, 8, f"{res['year']}年{res['month']}月分", align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(4)

            pdf.set_font("JP", "B", 13)
            pdf.cell(0, 8, f"氏名: {res['name']} 様", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("JP", "", 10)
            wage_label = "月給(基本給)" if res.get("pay_type") == "月給" else "時給"
            wage_value = (yen(res["monthly_base_salary"]) if res.get("pay_type") == "月給"
                          else yen(res["hourly_wage"]))
            pdf.cell(0, 7, f"給与形態: {wage_label}　{wage_label}: {wage_value}",
                     new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)

            bd = res["breakdown"]
            pdf.set_font("JP", "B", 11)
            pdf.set_fill_color(90, 90, 90)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 7, "勤怠", fill=True, new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("JP", "", 10)
            pdf.cell(0, 7,
                     f"総労働時間: {bd['total_hours']}h　"
                     f"残業25%: {bd['overtime25_hours']}h　残業50%: {bd['overtime50_hours']}h",
                     new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 7,
                     f"法定休日: {bd['holiday_hours']}h　深夜加算対象: {bd['night_hours']}h",
                     new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)

            gross = res["gross"]
            pay_items = [
                ("基本給",       yen(res["base_pay_amount"])),
                ("時給",         yen(res["hourly_pay_amount"])),
                ("残業25%",      yen(gross["pay_overtime25"])),
                ("残業50%",      yen(gross["pay_overtime50"])),
                ("休日出勤35%",  yen(gross["pay_holiday"])),
                ("深夜割増加算", yen(gross["pay_night_addon"])),
                ("役職手当",     yen(res["position_allowance"])),
                ("皆勤手当",     yen(res["attendance_allowance"])),
                ("住宅手当",     yen(res["housing_allowance"])),
                ("交通費",       yen(res["commute_allowance"])),
            ]
            deduction_items = [
                ("健康保険",       yen(res["health_insurance"])),
                ("厚生年金",       yen(res["pension"])),
                ("介護保険",       yen(res["care_insurance"])),
                ("子育て支援金",   yen(res["child_support_levy"])),
                ("雇用保険",       yen(res["employment_insurance"])),
                ("社会保険料(計)", yen(res["social_insurance_total"])),
                ("所得税",         yen(res["income_tax"])),
                ("住民税",         yen(res["resident_tax"])),
            ]
            two_col_section(pdf, "支給", pay_items, "控除", deduction_items)

            pdf.set_font("JP", "B", 11)
            pdf.set_fill_color(213, 232, 212)
            pdf.cell(88, 8, f"総支給額: {yen(res['gross_total'])}", border=1, align="C", fill=True)
            pdf.set_x(pdf.l_margin + 88 + 4)
            pdf.cell(88, 8, f"控除合計: {yen(res['total_deductions'])}", border=1, align="C", fill=True,
                     new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)

            pdf.set_font("JP", "", 10)
            adj_line = f"その他調整額: {yen(res['adjustment_amount'])}"
            if res.get("adjustment_note"):
                adj_line += f"（{res['adjustment_note']}）"
            pdf.cell(0, 7, adj_line, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)

            pdf.set_font("JP", "B", 16)
            pdf.set_fill_color(213, 232, 212)
            pdf.cell(0, 13, f"差引支給額: {yen(res['net_pay'])}", border=1, align="C", fill=True,
                     new_x="LMARGIN", new_y="NEXT")
            pdf.ln(6)

            pdf.set_font("JP", "", 8)
            pdf.set_text_color(120, 120, 120)
            pdf.multi_cell(
                0, 5,
                "※概算計算です。社会保険料は簡易計算（標準報酬月額等級表は未使用）、"
                "所得税(乙欄)は簡易近似です。交通費は非課税として所得税の課税対象額から除外しています。"
                "実際の給与支払前に社会保険労務士・税理士等の確認を必ず受けてください。",
            )

            safe_name = res["name"] or "無名"
            for ch in '\\/:*?"<>|':
                safe_name = safe_name.replace(ch, "")
            safe_name = safe_name or "無名"
            base_filename = f"給与明細_{year}-{month:02d}_{safe_name}"
            filename = base_filename
            n = 2
            while filename in used_names:
                filename = f"{base_filename}_{n}"
                n += 1
            used_names.add(filename)

            out_path = os.path.join(folder, filename + ".pdf")
            pdf.output(out_path)
            saved_paths.append(out_path)

        messagebox.showinfo(
            "完了",
            f"給与明細PDFを{len(saved_paths)}件、以下のフォルダに保存しました。\n{folder}",
        )

    # ── ログアウト ────────────────────────────────
    def on_logout(self):
        """ログアウト処理"""
        if messagebox.askyesno("ログアウト", "ログアウトしますか？"):
            self.destroy()


if __name__ == "__main__":
    _create_default_user()

    login = LoginWindow()
    login.mainloop()
    username = login.logged_in_user
    if username:
        app = AttendanceGUI(username=username)
        app.mainloop()
